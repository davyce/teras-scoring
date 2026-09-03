"""
Vues API Django REST Framework pour TERAS Entreprise
Gestion du dashboard, clients B2B, employés, documents, rapports, conformité

ENDPOINTS:
- GET  /api/enterprise/dashboard/              Dashboard principal
- GET  /api/enterprise/clients/                Liste clients
- POST /api/enterprise/clients/                Créer client
- GET  /api/enterprise/clients/{id}/           Détail client
- PATCH/DELETE /api/enterprise/clients/{id}/   Modifier/Supprimer
- GET  /api/enterprise/employees/              Liste employés
- POST /api/enterprise/employees/              Ajouter employé
- PATCH/DELETE /api/enterprise/employees/{id}/ Modifier/Supprimer
- GET  /api/enterprise/documents/              Liste documents
- POST /api/enterprise/documents/upload/       Upload document
- GET  /api/enterprise/compliance/             Statut conformité
- GET  /api/enterprise/reports/                Liste rapports
- POST /api/enterprise/reports/generate/       Générer rapport
- GET  /api/enterprise/analytics/sector/       Analytics sectorielle
- GET  /api/enterprise/analytics/trends/       Tendances
"""

import logging
import re
from pathlib import Path

from rest_framework import status, generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models import Q, Count, Avg, Sum
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal

from .document_parser import parse_document
from .models_enterprise import (
    EnterpriseClient,
    Employee,
    EnterpriseDocument,
    ComplianceStatus,
    EnterpriseReport,
    EnterpriseScore
)
from .serializers_enterprise import (
    EnterpriseClientSerializer,
    EnterpriseClientDetailSerializer,
    EnterpriseClientCreateSerializer,
    EmployeeSerializer,
    EmployeeCreateSerializer,
    EnterpriseDocumentSerializer,
    EnterpriseDocumentUploadSerializer,
    ComplianceStatusSerializer,
    EnterpriseReportSerializer,
    EnterpriseReportGenerateSerializer,
    EnterpriseDashboardSerializer,
    SectorAnalyticsSerializer,
    EnterpriseScoreSerializer,
)

logger = logging.getLogger(__name__)
ENTERPRISE_ANALYSIS_VERSION = 'local-enterprise-analysis-v2'


class IsEnterpriseUser(IsAuthenticated):
    """
    Permission personnalisée: utilisateur doit être authentifié ET de type 'enterprise'
    """
    def has_permission(self, request, view):
        is_authenticated = super().has_permission(request, view)
        if not is_authenticated:
            return False
        
        # Vérifier que l'utilisateur est bien une entreprise
        return request.user.user_type == 'enterprise'


def _safe_float(value, default: float = 0.0) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return default


def _label_enterprise_dossier_quality(completeness: float, analyzed_count: int, applied_count: int) -> str:
    if completeness >= 0.9 and analyzed_count >= 4 and applied_count >= 2:
        return 'robuste'
    if completeness >= 0.65 and analyzed_count >= 2:
        return 'exploitable'
    if analyzed_count >= 1:
        return 'partiel'
    return 'a_structurer'


def _normalize_enterprise_text(value) -> str:
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def _extract_enterprise_fcfa_amounts(text: str) -> list[float]:
    if not text:
        return []

    amounts: list[float] = []
    seen: set[int] = set()
    patterns = [
        r'(\d[\d\s.,]{2,})\s*(?:FCFA|XAF)',
        r'(?:FCFA|XAF)\s*(\d[\d\s.,]{2,})',
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            digits = re.sub(r'[^0-9]', '', match.group(1))
            if len(digits) < 4:
                continue
            value = int(digits)
            if value in seen:
                continue
            seen.add(value)
            amounts.append(float(value))
    return amounts


def _extract_labeled_fcfa_amount(text: str, labels: list[str]) -> float:
    if not text:
        return 0.0

    for label in labels:
        pattern = rf'(?:{label})[\s:;\-]*([\s\S]{{0,90}}?)'
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if not match:
            continue
        amounts = _extract_enterprise_fcfa_amounts(match.group(1))
        if amounts:
            return amounts[0]
    return 0.0


def _read_enterprise_document_text(document: EnterpriseDocument) -> str:
    if not document.file:
        return ''

    suffix = Path(document.file.name).suffix.lower()
    file_path = document.file.path

    try:
        if suffix == '.pdf':
            from pdfminer.high_level import extract_text

            return extract_text(file_path)[:20000]
        if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'}:
            from openpyxl import load_workbook

            workbook = load_workbook(file_path, data_only=True, read_only=True)
            rows: list[str] = []
            for worksheet in workbook.worksheets[:2]:
                for row in worksheet.iter_rows(min_row=1, max_row=40, values_only=True):
                    line = ' '.join(_normalize_enterprise_text(cell) for cell in row if _normalize_enterprise_text(cell))
                    if line:
                        rows.append(line)
            return '\n'.join(rows)[:20000]
        if suffix == '.csv':
            return Path(file_path).read_text(encoding='utf-8', errors='ignore')[:20000]
    except Exception as exc:
        logger.warning("Enterprise document text extraction failed for %s: %s", document.id, exc)

    return ''


def _extract_asset_register_totals(document: EnterpriseDocument, extracted_text: str) -> tuple[float, int]:
    if not document.file:
        return 0.0, 0

    suffix = Path(document.file.name).suffix.lower()
    if suffix not in {'.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'}:
        amounts = _extract_enterprise_fcfa_amounts(extracted_text)
        return float(sum(amounts)), len(amounts)

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(document.file.path, data_only=True, read_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(min_row=1, max_row=80, values_only=True))
        target_columns: set[int] = set()

        for row in rows[:5]:
            for index, cell in enumerate(row):
                label = _normalize_enterprise_text(cell).lower()
                if any(keyword in label for keyword in ('valeur', 'value', 'estime', 'estimated', 'montant', 'amount')):
                    target_columns.add(index)

        amounts: list[float] = []
        for row in rows[1:]:
            for index in target_columns:
                if index >= len(row):
                    continue
                value = _safe_float(row[index])
                if value >= 10_000:
                    amounts.append(value)

        if amounts:
            return float(sum(amounts)), len(amounts)
    except Exception as exc:
        logger.warning("Enterprise asset register extraction failed for %s: %s", document.id, exc)

    fallback_amounts = _extract_enterprise_fcfa_amounts(extracted_text)
    return float(sum(fallback_amounts)), len(fallback_amounts)


def _infer_enterprise_document_role(document: EnterpriseDocument, extracted_text: str) -> str:
    file_name = Path(document.file.name).name if document.file else ''
    raw = ' '.join(filter(None, [
        document.category,
        document.title,
        file_name,
        extracted_text[:1500],
    ])).lower()
    raw = raw.replace('-', '_')

    if (
        'fixed_assets_register' in raw
        or 'asset_inventory' in raw
        or 'registre des immobilisations' in raw
        or 'registre d immobilisations' in raw
    ):
        return 'asset_register'
    if 'asset_statement' in raw or 'etat des actifs' in raw:
        return 'asset_statement'
    if 'vehicle_registration' in raw or 'carte grise' in raw:
        return 'vehicle_title'
    if (
        'warehouse_lease_or_title' in raw
        or 'property_title' in raw
        or 'bail ou titre' in raw
        or "titre d'entrep" in raw
        or "titre d entrep" in raw
        or 'titre de propriete' in raw
        or 'bail commercial' in raw
    ):
        return 'property_or_lease'
    if document.category == 'invoice' or file_name.lower().startswith('invoice_') or 'facture' in raw:
        return 'invoice_evidence'
    if document.category == 'tax_filing':
        return 'tax_filing'
    if document.category == 'balance_sheet':
        return 'balance_sheet'
    if document.category == 'payroll':
        return 'payroll'
    if document.category == 'contract':
        return 'contract'
    if document.category == 'bank_statement':
        return 'bank_statement'
    return 'supporting'


def _enterprise_document_role_label(role: str) -> str:
    return {
        'asset_register': 'Registre des actifs',
        'asset_statement': 'État des actifs',
        'vehicle_title': 'Titre véhicule / carte grise',
        'property_or_lease': 'Titre ou bail immobilier',
        'invoice_evidence': 'Facture de vente',
        'tax_filing': 'Déclaration fiscale',
        'balance_sheet': 'Bilan comptable',
        'payroll': 'Fiche de paie',
        'contract': 'Contrat',
        'bank_statement': 'Relevé bancaire',
        'supporting': 'Pièce métier',
    }.get(role, 'Pièce métier')


def _enterprise_document_extraction_goal(role: str) -> str:
    return {
        'asset_register': 'inventaire des actifs et garanties mobilisables',
        'asset_statement': 'preuve des biens declares et valeur patrimoniale',
        'vehicle_title': 'preuve de propriete vehicule et garantie mobilisable',
        'property_or_lease': 'preuve immobiliere ou bail d exploitation',
        'invoice_evidence': 'preuve de chiffre d affaires et traction commerciale',
        'tax_filing': 'conformite fiscale et transparence administrative',
        'balance_sheet': 'solidite comptable et capacite financiere',
        'payroll': 'stabilite de la masse salariale et emploi local',
        'bank_statement': 'tresorerie observee et capacite de remboursement',
        'contract': 'visibilite commerciale et recurrence contractuelle',
        'supporting': 'preuve documentaire complementaire',
    }.get(role, 'preuve documentaire complementaire')


def _enterprise_document_needs_refresh(document: EnterpriseDocument) -> bool:
    summary = document.analysis_summary or {}
    if not summary:
        return True
    if summary.get('analysis_origin') != ENTERPRISE_ANALYSIS_VERSION:
        return True
    signals = summary.get('document_signals') or {}
    return not signals.get('document_role')


def _infer_enterprise_upload_category(requested_category: str, file_name: str, title: str) -> str:
    raw = f"{requested_category} {file_name} {title}".lower().replace('-', '_')
    asset_like = any(token in raw for token in (
        'asset_statement',
        'fixed_assets',
        'asset_inventory',
        'vehicle_registration',
        'warehouse_lease_or_title',
        'property_title',
        'carte grise',
        'titre de propriete',
    ))
    if 'invoice_' in raw or 'facture' in raw:
        return 'invoice'
    if asset_like:
        return requested_category if requested_category in {'balance_sheet', 'other'} else 'other'
    if any(token in raw for token in ('tax', 'fiscal')):
        return 'tax_filing'
    if any(token in raw for token in ('payroll', 'paie', 'salary')):
        return 'payroll'
    if any(token in raw for token in ('contract', 'contrat')):
        return 'contract'
    if any(token in raw for token in ('bank_statement', 'releve', 'statement')):
        return 'bank_statement'
    return requested_category or 'other'


def _build_enterprise_document_analysis(document: EnterpriseDocument) -> dict:
    extracted_text = _read_enterprise_document_text(document)
    category = document.category
    document_role = _infer_enterprise_document_role(document, extracted_text)
    parse_expected = document_role in {'bank_statement', 'invoice_evidence'}
    parser_result = None
    extraction_goal = _enterprise_document_extraction_goal(document_role)

    if parse_expected and document.file:
        try:
            parser_result = parse_document(document.file.path, Path(document.file.name).name)
        except Exception as exc:
            logger.warning("Enterprise document parsing fallback for %s: %s", document.id, exc)

    extracted_metrics = {
        'monthly_revenue_xaf': 0,
        'monthly_cashflow_xaf': 0,
        'crm_estimated_xaf': 0,
        'transactions_count': 0,
        'months_covered': 0,
        'authenticity_score': 0,
        'asset_value_xaf': 0,
        'asset_items_count': 0,
        'invoice_amount_xaf': 0,
        'invoice_count': 0,
        'collateral_value_xaf': 0,
    }
    strengths = []
    risks = []
    recommended_actions = []
    pillar_hints = {'T': 0.18, 'E': 0.18, 'R': 0.18, 'A': 0.18, 'S': 0.18}
    proof_modes: list[str] = []
    asset_proof_types: list[str] = []
    collateral_eligible = False

    if parser_result and parser_result.get('parsing_success'):
        signals = parser_result.get('teras_signals', {}) or {}
        quality = parser_result.get('quality', {}) or {}
        stats = quality.get('stats', {}) or {}
        extracted_metrics.update({
            'monthly_revenue_xaf': round(_safe_float(signals.get('income_signal', {}).get('monthly_avg_xaf'))),
            'monthly_cashflow_xaf': round(_safe_float(stats.get('net_cashflow_xaf'))),
            'crm_estimated_xaf': round(_safe_float(signals.get('crm_estimated_xaf'))),
            'transactions_count': int(stats.get('total_transactions') or len(parser_result.get('transactions', []))),
            'months_covered': int(stats.get('months_covered') or signals.get('months_analyzed') or 0),
            'authenticity_score': round(_safe_float(quality.get('authenticity_score')), 3),
        })
        strengths.append("des flux financiers exploitables ont ete extraits du document")
        if extracted_metrics['crm_estimated_xaf'] > 0:
            strengths.append(f"capacite de remboursement estimee a {extracted_metrics['crm_estimated_xaf']:,} FCFA/mois")
        if extracted_metrics['authenticity_score'] >= 0.75:
            strengths.append("le document parait suffisamment fiable pour nourrir TERAS")
        if extracted_metrics['monthly_cashflow_xaf'] < 0:
            risks.append("cashflow net negatif detecte sur la periode analysee")
        pillar_hints['T'] += min(0.28, extracted_metrics['authenticity_score'] * 0.22)
        pillar_hints['A'] += min(0.32, extracted_metrics['months_covered'] * 0.04)
        pillar_hints['R'] += min(0.24, (_safe_float(signals.get('income_signal', {}).get('income_stability')) * 0.24))
        if document_role == 'invoice_evidence' and extracted_metrics['monthly_revenue_xaf'] <= 0:
            extracted_metrics['invoice_amount_xaf'] = round(
                _extract_labeled_fcfa_amount(
                    extracted_text,
                    ['Montant TTC', 'Total TTC', 'Montant', 'Total'],
                )
            )
            if extracted_metrics['invoice_amount_xaf'] > 0:
                extracted_metrics['invoice_count'] = 1
                extracted_metrics['monthly_revenue_xaf'] = extracted_metrics['invoice_amount_xaf']
                strengths.append(
                    f"montant facture objectivé a {extracted_metrics['invoice_amount_xaf']:,} FCFA"
                )
                recommended_actions.append("associer cette facture a son relevé bancaire pour confirmer l'encaissement")
                pillar_hints['R'] += 0.18
                pillar_hints['A'] += 0.12

    if document_role in {'asset_statement', 'asset_register', 'vehicle_title', 'property_or_lease'}:
        if document_role == 'asset_register':
            asset_total, asset_count = _extract_asset_register_totals(document, extracted_text)
        else:
            asset_amounts = _extract_enterprise_fcfa_amounts(extracted_text)
            if document_role in {'vehicle_title', 'property_or_lease'} and asset_amounts:
                asset_amounts = asset_amounts[:1]
            asset_total = float(sum(asset_amounts))
            asset_count = len(asset_amounts) if asset_amounts else 0

        if asset_total > 0:
            extracted_metrics['asset_value_xaf'] = round(asset_total)
            extracted_metrics['asset_items_count'] = max(asset_count, 1)

            for token in ('vehicle_registration', 'equipment_purchase_invoice', 'property_title', 'warehouse_lease_or_title', 'lease_contract'):
                if token in extracted_text.lower():
                    proof_modes.append(token)
            if document_role == 'vehicle_title':
                proof_modes.append('vehicle_registration')
                asset_proof_types.append('vehicule')
                collateral_eligible = True
                extracted_metrics['collateral_value_xaf'] = round(asset_total * 0.65)
            elif document_role == 'property_or_lease':
                asset_proof_types.append('immobilier')
                coefficient = 0.8 if 'titre' in extracted_text.lower() or 'title' in extracted_text.lower() else 0.4
                collateral_eligible = True
                extracted_metrics['collateral_value_xaf'] = round(asset_total * coefficient)
            elif document_role in {'asset_statement', 'asset_register'}:
                if not asset_proof_types:
                    asset_proof_types.append('equipement')
                collateral_eligible = True
                extracted_metrics['collateral_value_xaf'] = round(asset_total * 0.5)

            strengths.append(
                f"preuve d'actifs documentee pour {extracted_metrics['asset_items_count']} element(s), valeur estimée {round(asset_total):,} FCFA"
            )
            pillar_hints['A'] += min(0.34, 0.16 + min(asset_total / 20_000_000, 1.0) * 0.22)
            pillar_hints['T'] += 0.10
            recommended_actions.append("maintenir les preuves de propriété et les valeurs d'actifs à jour dans le dossier")
        else:
            risks.append("la piece d'actif ne contient pas encore de valeur exploitable")

    elif document_role == 'invoice_evidence' and extracted_metrics['invoice_amount_xaf'] <= 0:
        invoice_amount = _extract_labeled_fcfa_amount(
            extracted_text,
            ['Montant TTC', 'Total TTC', 'Montant', 'Total'],
        )
        if not invoice_amount:
            amounts = _extract_enterprise_fcfa_amounts(extracted_text)
            invoice_amount = amounts[0] if amounts else 0
        if invoice_amount > 0:
            extracted_metrics['invoice_amount_xaf'] = round(invoice_amount)
            extracted_metrics['invoice_count'] = 1
            extracted_metrics['monthly_revenue_xaf'] = max(
                extracted_metrics['monthly_revenue_xaf'],
                extracted_metrics['invoice_amount_xaf'],
            )
            strengths.append(f"facture exploitable detectee pour {round(invoice_amount):,} FCFA")
            recommended_actions.append("associer cette facture à un contrat ou un relevé d'encaissement pour un impact plus fort")
            pillar_hints['R'] += 0.20
            pillar_hints['A'] += 0.14
        else:
            risks.append("la facture n'a pas livré de montant TTC exploitable")

    if not strengths:
        if document_role == 'tax_filing':
            strengths.append("piece fiscale disponible pour appuyer la transparence de l'entreprise")
            pillar_hints['T'] += 0.42
            recommended_actions.append("ajouter la preuve de depot ou le recu fiscal correspondant")
        elif document_role == 'balance_sheet':
            strengths.append("bilan comptable disponible pour appuyer les capacites financieres")
            pillar_hints['T'] += 0.18
            pillar_hints['A'] += 0.28
            recommended_actions.append("completer avec un relevé bancaire recent pour objectiver la tresorerie")
        elif document_role == 'payroll':
            strengths.append("document de paie disponible pour objectiver l'emploi et la masse salariale")
            pillar_hints['E'] += 0.38
            pillar_hints['S'] += 0.16
            recommended_actions.append("maintenir des periodes de paie regulieres pour consolider le pilier E")
        elif document_role == 'invoice_evidence':
            strengths.append("piece commerciale disponible pour attester l'activite")
            pillar_hints['A'] += 0.24
            pillar_hints['R'] += 0.22
            recommended_actions.append("regrouper plusieurs factures ou un registre de ventes pour gagner en profondeur")
        elif document_role == 'contract':
            strengths.append("contrat disponible pour attester la stabilite des relations commerciales")
            pillar_hints['R'] += 0.26
            pillar_hints['S'] += 0.18
            recommended_actions.append("ajouter les pieces d'execution du contrat pour renforcer le dossier")
        else:
            strengths.append("document metier archivé dans le dossier entreprise")
            recommended_actions.append("associer ce document a une piece financiere ou fiscale pour un impact TERAS plus fort")

    if not document.period:
        risks.append("periode documentaire non renseignee")
    if document.status == 'rejected':
        risks.append("document deja marque comme rejete ou incomplet")
    if not recommended_actions:
        recommended_actions.append("completer le dossier avec des pieces couvrant fiscalite, activite et tresorerie")

    narrative = (
        "## Synthese\n"
        f"{document.title} ({_enterprise_document_role_label(document_role)}) a ete relu pour l'entreprise. "
        f"Cette piece nourrit surtout les piliers {', '.join(code for code, value in pillar_hints.items() if value >= 0.3)}.\n\n"
        "## Points forts\n"
        + '\n'.join(f"- {item}" for item in strengths)
        + "\n\n## Points de vigilance\n"
        + '\n'.join(f"- {item}" for item in (risks or ['aucun risque majeur automatiquement detecte']))
        + "\n\n## Actions recommandees\n"
        + '\n'.join(f"- {item}" for item in recommended_actions[:3])
    )

    estimated_change = int(sum(max(0, hint - 0.18) for hint in pillar_hints.values()) * 180)
    return {
        'analysis_origin': ENTERPRISE_ANALYSIS_VERSION,
        'document_id': document.id,
        'document_title': document.title,
        'document_category': category,
        'document_role': document_role,
        'document_role_label': _enterprise_document_role_label(document_role),
        'category_display': document.get_category_display(),
        'period': document.period,
        'parse_expected': parse_expected,
        'extraction_goal': extraction_goal,
        'parser_result': parser_result if parser_result and parser_result.get('parsing_success') else None,
        'extracted_metrics': extracted_metrics,
        'strengths': strengths,
        'risks': risks,
        'recommended_actions': recommended_actions,
        'score_impact': {
            'pillar_hints': {key: round(min(1.0, max(0.0, value)), 3) for key, value in pillar_hints.items()},
            'estimated_change': estimated_change,
        },
        'document_signals': {
            'document_quality': 'asset' if document_role in {'asset_statement', 'asset_register', 'vehicle_title', 'property_or_lease'} else ('financial' if document_role in {'bank_statement', 'balance_sheet', 'invoice_evidence'} else 'supporting'),
            'evidence_strength': 'high' if parse_expected or document_role in {'tax_filing', 'contract', 'vehicle_title', 'property_or_lease'} else 'medium',
            'requires_pairing': document_role in {'invoice_evidence', 'contract', 'balance_sheet'},
            'suggested_pairs': {
                'invoice_evidence': ['bank_statement', 'contract'],
                'contract': ['invoice', 'bank_statement'],
                'balance_sheet': ['bank_statement', 'tax_filing'],
                'asset_statement': ['fixed_assets_register', 'invoice'],
                'asset_register': ['bank_statement', 'contract'],
                'vehicle_title': ['asset_statement'],
                'property_or_lease': ['contract', 'bank_statement'],
            }.get(document_role, []),
            'document_role': document_role,
            'document_role_label': _enterprise_document_role_label(document_role),
            'asset_proof_types': list(dict.fromkeys(asset_proof_types)),
            'proof_modes': list(dict.fromkeys(proof_modes)),
            'collateral_eligible': collateral_eligible,
        },
        'analysis_text': narrative,
        'analyzed_at': timezone.now().isoformat(),
    }


def _recompute_enterprise_from_documents(enterprise, current_document=None):
    documents = list(EnterpriseDocument.objects.filter(enterprise=enterprise).order_by('-uploaded_at'))
    current_document_id = getattr(current_document, 'id', None)
    analyzed_documents = [
        doc
        for doc in documents
        if (
            doc.analysis_summary
            and not _enterprise_document_needs_refresh(doc)
            and (
                (doc.analysis_summary or {}).get('applied_to_teras')
                or doc.id == current_document_id
            )
        )
    ]
    categories = set()
    applied_documents = [
        doc for doc in analyzed_documents
        if (doc.analysis_summary or {}).get('applied_to_teras') or doc.id == current_document_id
    ]

    total_docs = len(documents)
    completeness = min(1.0, len(analyzed_documents) / 6) if analyzed_documents else 0.0

    total_employees = Employee.objects.filter(enterprise=enterprise, status='active').count()
    local_employees = Employee.objects.filter(enterprise=enterprise, status='active', is_local=True).count()
    local_ratio = (local_employees / total_employees) if total_employees else 0.0
    active_clients = EnterpriseClient.objects.filter(enterprise=enterprise, status='active').count()

    monthly_revenues = []
    monthly_cashflows = []
    authenticity_scores = []
    assets_documented_total = 0.0
    assets_verified_count = 0
    invoice_amount_total = 0.0
    invoices_analyzed_count = 0
    collateral_value_total = 0.0
    asset_proof_types: set[str] = set()
    latest_summary = None
    latest_processed_doc = next((doc for doc in analyzed_documents if doc.processed_at), None)
    for doc in analyzed_documents:
        summary = doc.analysis_summary or {}
        latest_summary = latest_summary or summary
        signals = summary.get('document_signals', {}) or {}
        role = signals.get('document_role') or summary.get('document_role') or doc.category
        categories.add(role)
        metrics = summary.get('extracted_metrics', {})
        revenue = _safe_float(metrics.get('monthly_revenue_xaf'))
        cashflow = _safe_float(metrics.get('monthly_cashflow_xaf'))
        auth = _safe_float(metrics.get('authenticity_score'))
        assets_documented_total += _safe_float(metrics.get('asset_value_xaf'))
        assets_verified_count += int(metrics.get('asset_items_count') or (1 if _safe_float(metrics.get('asset_value_xaf')) > 0 else 0))
        invoice_amount_total += _safe_float(metrics.get('invoice_amount_xaf'))
        collateral_value_total += _safe_float(metrics.get('collateral_value_xaf'))
        asset_proof_types.update(signals.get('asset_proof_types') or [])
        if _safe_float(metrics.get('invoice_amount_xaf')) > 0 or role == 'invoice_evidence':
            invoices_analyzed_count += int(metrics.get('invoice_count') or 1)
        if revenue > 0:
            monthly_revenues.append(revenue)
        monthly_cashflows.append(cashflow)
        if auth > 0:
            authenticity_scores.append(auth)

    has_tax = 'tax_filing' in categories
    has_balance = 'balance_sheet' in categories or 'asset_statement' in categories or 'asset_register' in categories
    has_payroll = 'payroll' in categories
    has_bank = 'bank_statement' in categories
    has_invoice = 'invoice_evidence' in categories or 'invoice' in categories
    has_contract = 'contract' in categories

    avg_revenue = sum(monthly_revenues) / len(monthly_revenues) if monthly_revenues else 0.0
    avg_cashflow = sum(monthly_cashflows) / len(monthly_cashflows) if monthly_cashflows else 0.0
    avg_authenticity = sum(authenticity_scores) / len(authenticity_scores) if authenticity_scores else 0.0
    dossier_quality = _label_enterprise_dossier_quality(completeness, len(analyzed_documents), len(applied_documents))
    asset_factor = min(assets_documented_total / 20_000_000, 1.0) if assets_documented_total > 0 else 0.0
    collateral_factor = min(collateral_value_total / 15_000_000, 1.0) if collateral_value_total > 0 else 0.0
    invoice_factor = min(invoice_amount_total / 6_000_000, 1.0) if invoice_amount_total > 0 else 0.0

    T = min(1.0, 0.10 + (0.34 if has_tax else 0) + (0.18 if has_balance else 0) + (0.12 if has_bank else 0) + completeness * 0.18 + avg_authenticity * 0.08)
    E = min(1.0, 0.08 + (0.28 if has_payroll else 0) + local_ratio * 0.34 + min(total_employees / 20, 1.0) * 0.18 + completeness * 0.12)
    R = min(1.0, 0.10 + (0.24 if has_contract else 0) + (0.24 if has_invoice else 0) + min(active_clients / 15, 1.0) * 0.22 + completeness * 0.12 + invoice_factor * 0.10)
    revenue_factor = min(avg_revenue / 3_000_000, 1.0) if avg_revenue > 0 else 0.0
    cashflow_factor = 0.15 if avg_cashflow > 0 else 0.0
    A = min(1.0, 0.10 + (0.26 if has_bank else 0) + (0.18 if has_balance else 0) + (0.14 if has_invoice else 0) + revenue_factor * 0.20 + cashflow_factor + completeness * 0.10 + asset_factor * 0.18 + collateral_factor * 0.14)
    S = min(1.0, 0.12 + (0.18 if has_payroll else 0) + (0.16 if has_contract else 0) + (0.12 if has_tax else 0) + completeness * 0.18 + (0.12 if avg_cashflow >= 0 and total_docs else 0))

    breakdown = {key: round(value, 4) for key, value in {'T': T, 'E': E, 'R': R, 'A': A, 'S': S}.items()}
    score = int(round((T * 0.30 + E * 0.25 + R * 0.15 + A * 0.20 + S * 0.10) * 1000))

    compliance_status, _ = ComplianceStatus.objects.get_or_create(
        enterprise=enterprise,
        defaults={'compliance_rate': Decimal('0.00')},
    )

    missing_declarations = []
    active_alerts = []
    recommendations = []
    if not has_tax:
        missing_declarations.append('Déclaration fiscale recente')
        active_alerts.append('Aucune declaration fiscale recente analysee')
        recommendations.append('Ajouter une declaration fiscale recente pour renforcer le pilier T')
    if not has_bank:
        active_alerts.append('Aucun releve bancaire analyse pour objectiver la tresorerie')
        recommendations.append('Ajouter un releve bancaire recent pour mieux qualifier le pilier A')
    if not has_payroll and total_employees > 0:
        active_alerts.append('Aucune piece de paie disponible pour le pilier Emploi')
        recommendations.append('Ajouter au moins un document de paie pour consolider le pilier E')
    if assets_documented_total <= 0:
        recommendations.append("Ajouter des preuves d'actifs exploitables (carte grise, titre, registre d'immobilisations)")
    if invoice_amount_total <= 0:
        recommendations.append("Ajouter des factures ou bordereaux de ventes pour objectiver l'activite commerciale")
    if completeness < 0.5:
        recommendations.append('Completer le dossier avec au moins 4 categories documentaires distinctes')

    compliance_status.missing_declarations = missing_declarations
    compliance_status.active_alerts = active_alerts
    compliance_status.recommendations = recommendations
    compliance_status.late_payments = 1 if avg_cashflow < 0 and has_bank else 0
    compliance_status.penalties = Decimal('0.00')
    tax_documents = [doc for doc in documents if doc.category == 'tax_filing' and doc.period_end]
    latest_tax_document = max(tax_documents, key=lambda doc: doc.period_end) if tax_documents else None
    if latest_tax_document and latest_tax_document.period_end:
        compliance_status.last_tax_filing = latest_tax_document.period_end
    compliance_status.last_audit_date = timezone.now().date()
    compliance_status.next_audit_date = timezone.now().date() + timedelta(days=90)
    compliance_status.save()

    enterprise_score = EnterpriseScore.objects.create(
        enterprise=enterprise,
        score=max(0, min(1000, score)),
        breakdown=breakdown,
        input_data={
            'documents_total': total_docs,
            'documents_analyzed': len(analyzed_documents),
            'documents_applied': len(applied_documents),
            'categories': sorted(categories),
            'avg_monthly_revenue_xaf': round(avg_revenue),
            'avg_monthly_cashflow_xaf': round(avg_cashflow),
            'avg_authenticity': round(avg_authenticity, 3),
            'assets_documented_total_xaf': round(assets_documented_total),
            'assets_verified_count': assets_verified_count,
            'invoice_amount_total_xaf': round(invoice_amount_total),
            'invoices_analyzed_count': invoices_analyzed_count,
            'collateral_value_xaf': round(collateral_value_total),
            'collateral_strength': 'high' if collateral_factor >= 0.65 else ('medium' if collateral_factor >= 0.3 else 'low'),
            'asset_proof_types': sorted(asset_proof_types),
            'active_clients': active_clients,
            'employees_total': total_employees,
            'local_employees_ratio': round(local_ratio, 3),
            'completeness_ratio': round(completeness, 3),
            'dossier_quality': dossier_quality,
            'latest_processed_at': latest_processed_doc.processed_at.isoformat() if latest_processed_doc and latest_processed_doc.processed_at else None,
            'alerts': active_alerts,
            'latest_summary': latest_summary,
        },
        sector='enterprise-documents',
    )

    return enterprise_score, {
        'documents_total': total_docs,
        'documents_analyzed': len(analyzed_documents),
        'documents_applied': len(applied_documents),
        'categories': sorted(categories),
        'avg_monthly_revenue_xaf': round(avg_revenue),
        'avg_monthly_cashflow_xaf': round(avg_cashflow),
        'avg_authenticity': round(avg_authenticity, 3),
        'assets_documented_total_xaf': round(assets_documented_total),
        'assets_verified_count': assets_verified_count,
        'invoice_amount_total_xaf': round(invoice_amount_total),
        'invoices_analyzed_count': invoices_analyzed_count,
        'collateral_value_xaf': round(collateral_value_total),
        'collateral_strength': 'high' if collateral_factor >= 0.65 else ('medium' if collateral_factor >= 0.3 else 'low'),
        'asset_proof_types': sorted(asset_proof_types),
        'completeness_ratio': round(completeness, 3),
        'dossier_quality': dossier_quality,
        'latest_processed_at': latest_processed_doc.processed_at.isoformat() if latest_processed_doc and latest_processed_doc.processed_at else None,
        'latest_summary': latest_summary,
        'alerts': active_alerts,
    }


# ============================================================================
# DASHBOARD ENTERPRISE
# ============================================================================

class EnterpriseDashboardView(APIView):
    """
    GET /api/enterprise/dashboard/
    
    Retourne le dashboard complet de l'entreprise:
    - Score TERAS Entreprise actuel
    - Breakdown des 5 piliers (T, E, R, A, S)
    - KPIs: clients, employés, conformité
    - Historique des scores (12 derniers mois)
    - Comparaison sectorielle
    - Alertes actives
    - Recommandations IA
    """
    permission_classes = [IsEnterpriseUser]
    
    def get(self, request):
        try:
            enterprise = request.user
            
            # 1. Score TERAS actuel
            latest_score = EnterpriseScore.objects.filter(
                enterprise=enterprise
            ).order_by('-computed_at').first()
            
            current_score = latest_score.score if latest_score else 0
            breakdown = latest_score.breakdown if latest_score else {
                'T': 0.0, 'E': 0.0, 'R': 0.0, 'A': 0.0, 'S': 0.0
            }
            
            # 2. Calculer la tendance (comparer avec le score précédent)
            previous_score = EnterpriseScore.objects.filter(
                enterprise=enterprise
            ).order_by('-computed_at')[1:2].first()
            
            if previous_score:
                score_change = current_score - previous_score.score
                score_trend = 'up' if score_change > 0 else ('down' if score_change < 0 else 'stable')
            else:
                score_change = 0
                score_trend = 'stable'
            
            # 3. KPIs Clients
            total_clients = EnterpriseClient.objects.filter(enterprise=enterprise).count()
            active_clients = EnterpriseClient.objects.filter(
                enterprise=enterprise,
                status='active'
            ).count()
            
            # 4. KPIs Employés
            total_employees = Employee.objects.filter(
                enterprise=enterprise,
                status='active'
            ).count()
            local_employees = Employee.objects.filter(
                enterprise=enterprise,
                status='active',
                is_local=True
            ).count()
            
            # 5. Conformité
            compliance_status, _ = ComplianceStatus.objects.get_or_create(
                enterprise=enterprise,
                defaults={'compliance_rate': Decimal('0.00')}
            )
            compliance_rate = compliance_status.compliance_rate

            documents_qs = EnterpriseDocument.objects.filter(enterprise=enterprise).order_by('-processed_at', '-uploaded_at')
            documents_total = documents_qs.count()
            documents_analyzed = documents_qs.exclude(analysis_summary__isnull=True).exclude(analysis_summary={}).count()
            documents_applied = sum(
                1 for doc in documents_qs
                if (doc.analysis_summary or {}).get('applied_to_teras')
            )
            latest_document = next((doc for doc in documents_qs if doc.analysis_summary), None)
            latest_input_data = latest_score.input_data if latest_score and latest_score.input_data else {}
            categories = latest_input_data.get('categories') or sorted({doc.category for doc in documents_qs})
            document_intelligence = {
                'documents_total': documents_total,
                'documents_analyzed': documents_analyzed,
                'documents_applied': documents_applied,
                'categories': categories,
                'completeness_ratio': latest_input_data.get('completeness_ratio', round(min(1.0, documents_total / 6), 3) if documents_total else 0.0),
                'avg_monthly_revenue_xaf': latest_input_data.get('avg_monthly_revenue_xaf', 0),
                'avg_monthly_cashflow_xaf': latest_input_data.get('avg_monthly_cashflow_xaf', 0),
                'avg_authenticity': latest_input_data.get('avg_authenticity', 0),
                'assets_documented_total_xaf': latest_input_data.get('assets_documented_total_xaf', 0),
                'assets_verified_count': latest_input_data.get('assets_verified_count', 0),
                'invoice_amount_total_xaf': latest_input_data.get('invoice_amount_total_xaf', 0),
                'invoices_analyzed_count': latest_input_data.get('invoices_analyzed_count', 0),
                'collateral_value_xaf': latest_input_data.get('collateral_value_xaf', 0),
                'collateral_strength': latest_input_data.get('collateral_strength', 'low'),
                'asset_proof_types': latest_input_data.get('asset_proof_types', []),
                'dossier_quality': latest_input_data.get('dossier_quality', _label_enterprise_dossier_quality(
                    latest_input_data.get('completeness_ratio', round(min(1.0, documents_total / 6), 3) if documents_total else 0.0),
                    documents_analyzed,
                    documents_applied,
                )),
                'latest_processed_at': latest_input_data.get(
                    'latest_processed_at',
                    latest_document.processed_at.isoformat() if latest_document and latest_document.processed_at else None,
                ),
                'latest_summary': latest_input_data.get(
                    'latest_summary',
                    latest_document.analysis_summary if latest_document else None,
                ),
                'alerts': latest_input_data.get('alerts', compliance_status.active_alerts if compliance_status else []),
            }
            
            # 6. Historique des scores (12 derniers mois)
            twelve_months_ago = timezone.now() - timedelta(days=365)
            score_history = EnterpriseScore.objects.filter(
                enterprise=enterprise,
                computed_at__gte=twelve_months_ago
            ).order_by('computed_at')[:12]
            
            score_history_data = EnterpriseScoreSerializer(
                score_history,
                many=True
            ).data
            
            # 7. Comparaison sectorielle
            sector = latest_score.sector if latest_score else 'Non défini'
            sector_average = latest_score.sector_average if latest_score else None
            percentile = latest_score.percentile if latest_score else None
            
            sector_comparison = {
                'sector': sector,
                'your_score': current_score,
                'sector_average': sector_average,
                'percentile': percentile,
                'above_average': current_score > sector_average if sector_average else None
            }
            
            # 8. Alertes actives
            active_alerts = compliance_status.active_alerts if compliance_status else []
            
            # 9. Recommandations IA
            recommendations = self._generate_recommendations(
                breakdown,
                current_score,
                compliance_rate,
                total_employees,
                local_employees
            )
            
            # 10. Construire la réponse
            dashboard_data = {
                'current_score': current_score,
                'score_trend': score_trend,
                'score_change': score_change,
                'breakdown': breakdown,
                'total_clients': total_clients,
                'active_clients': active_clients,
                'total_employees': total_employees,
                'local_employees': local_employees,
                'compliance_rate': float(compliance_rate),
                'score_history': score_history_data,
                'sector_comparison': sector_comparison,
                'active_alerts': active_alerts,
                'recommendations': recommendations,
                'document_intelligence': document_intelligence,
            }
            
            return Response(dashboard_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Erreur lors du chargement du dashboard: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _generate_recommendations(self, breakdown, score, compliance_rate, total_emp, local_emp):
        """
        Génère des recommandations personnalisées basées sur les piliers faibles
        """
        recommendations = []
        
        # Recommandations par pilier (seuils indicatifs sur 1.0)
        # T - Transparence (30% = 300 points max)
        if breakdown.get('T', 0) < 0.70:  # Moins de 70% du max
            recommendations.append({
                'pillar': 'T',
                'title': 'Améliorer la transparence fiscale',
                'description': 'Uploadez vos dernières déclarations fiscales et bilans comptables.',
                'impact': '+50 points',
                'priority': 'high'
            })
        
        # E - Emploi local (25% = 250 points max)
        if breakdown.get('E', 0) < 0.70:
            local_rate = (local_emp / total_emp * 100) if total_emp > 0 else 0
            recommendations.append({
                'pillar': 'E',
                'title': 'Augmenter l\'emploi local',
                'description': f'Votre taux d\'emploi local est de {local_rate:.1f}%. Visez 80%+.',
                'impact': '+40 points',
                'priority': 'high'
            })
        
        # R - Rétention (15% = 150 points max)
        if breakdown.get('R', 0) < 0.70:
            recommendations.append({
                'pillar': 'R',
                'title': 'Fidéliser vos clients',
                'description': 'Mettez en place des programmes de fidélité et suivez le NPS.',
                'impact': '+25 points',
                'priority': 'medium'
            })
        
        # A - Activité (20% = 200 points max)
        if breakdown.get('A', 0) < 0.70:
            recommendations.append({
                'pillar': 'A',
                'title': 'Diversifier votre activité',
                'description': 'Augmentez la fréquence de vos transactions et diversifiez votre clientèle.',
                'impact': '+35 points',
                'priority': 'medium'
            })
        
        # S - Stabilité sociale (10% = 100 points max)
        if breakdown.get('S', 0) < 0.70:
            recommendations.append({
                'pillar': 'S',
                'title': 'Renforcer la stabilité',
                'description': 'Réduisez les litiges et respectez les délais de paiement.',
                'impact': '+15 points',
                'priority': 'low'
            })
        
        # Recommandation conformité si < 80%
        if compliance_rate < 80:
            recommendations.insert(0, {
                'pillar': 'T',
                'title': 'URGENT: Régulariser la conformité',
                'description': f'Votre conformité est à {compliance_rate}%. Cela impacte fortement votre score.',
                'impact': '+60 points',
                'priority': 'critical'
            })
        
        return recommendations[:5]  # Max 5 recommandations


# ============================================================================
# CLIENTS B2B (PORTEFEUILLE)
# ============================================================================

class EnterpriseClientsListView(generics.ListCreateAPIView):
    """
    GET /api/enterprise/clients/
    Liste tous les clients de l'entreprise avec pagination
    
    POST /api/enterprise/clients/
    Créer un nouveau dossier client
    
    Query params:
    - search: Recherche par nom
    - client_type: Filtrer par type (individual, pme, company)
    - status: Filtrer par statut (active, pending, archived)
    - risk_level: Filtrer par risque (low, medium, high)
    - score_min / score_max: Filtrer par plage de score
    - ordering: Tri (score, -score, created_at, -created_at)
    """
    permission_classes = [IsEnterpriseUser]
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return EnterpriseClientCreateSerializer
        return EnterpriseClientSerializer
    
    def get_queryset(self):
        queryset = EnterpriseClient.objects.filter(
            enterprise=self.request.user
        )
        
        # Filtres
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(kyc_id__icontains=search) |
                Q(internal_ref__icontains=search)
            )
        
        client_type = self.request.query_params.get('client_type')
        if client_type:
            queryset = queryset.filter(client_type=client_type)
        
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        risk_level = self.request.query_params.get('risk_level')
        if risk_level:
            queryset = queryset.filter(risk_level=risk_level)
        
        score_min = self.request.query_params.get('score_min')
        if score_min:
            queryset = queryset.filter(teras_score__gte=int(score_min))
        
        score_max = self.request.query_params.get('score_max')
        if score_max:
            queryset = queryset.filter(teras_score__lte=int(score_max))
        
        # Tri
        ordering = self.request.query_params.get('ordering', '-created_at')
        queryset = queryset.order_by(ordering)
        
        return queryset


class EnterpriseClientDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    GET /api/enterprise/clients/{id}/
    Récupérer les détails complets d'un client
    
    PATCH /api/enterprise/clients/{id}/
    Mettre à jour un client
    
    DELETE /api/enterprise/clients/{id}/
    Supprimer un client
    """
    permission_classes = [IsEnterpriseUser]
    serializer_class = EnterpriseClientDetailSerializer
    
    def get_queryset(self):
        # Seulement les clients de l'entreprise connectée
        return EnterpriseClient.objects.filter(enterprise=self.request.user)


# ============================================================================
# EMPLOYÉS
# ============================================================================

class EnterpriseEmployeesListView(generics.ListCreateAPIView):
    """
    GET /api/enterprise/employees/
    Liste tous les employés de l'entreprise
    
    POST /api/enterprise/employees/
    Ajouter un nouvel employé
    
    Query params:
    - search: Recherche par nom
    - status: Filtrer par statut (active, on_leave, terminated)
    - is_local: Filtrer employés locaux (true/false)
    - employment_type: Filtrer par type de contrat
    """
    permission_classes = [IsEnterpriseUser]
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return EmployeeCreateSerializer
        return EmployeeSerializer
    
    def get_queryset(self):
        queryset = Employee.objects.filter(
            enterprise=self.request.user
        )
        
        # Filtres
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(employee_id__icontains=search) |
                Q(position__icontains=search)
            )
        
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        is_local = self.request.query_params.get('is_local')
        if is_local:
            queryset = queryset.filter(is_local=is_local.lower() == 'true')
        
        employment_type = self.request.query_params.get('employment_type')
        if employment_type:
            queryset = queryset.filter(employment_type=employment_type)
        
        return queryset.order_by('-hire_date')


class EnterpriseEmployeeDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    GET /api/enterprise/employees/{id}/
    Récupérer un employé
    
    PATCH /api/enterprise/employees/{id}/
    Modifier un employé
    
    DELETE /api/enterprise/employees/{id}/
    Supprimer un employé
    """
    permission_classes = [IsEnterpriseUser]
    serializer_class = EmployeeSerializer
    
    def get_queryset(self):
        return Employee.objects.filter(enterprise=self.request.user)


# ============================================================================
# DOCUMENTS
# ============================================================================

class EnterpriseDocumentsListView(generics.ListAPIView):
    """
    GET /api/enterprise/documents/
    Liste tous les documents uploadés
    
    Query params:
    - category: Filtrer par catégorie
    - status: Filtrer par statut
    - period: Filtrer par période
    """
    permission_classes = [IsEnterpriseUser]
    serializer_class = EnterpriseDocumentSerializer
    
    def get_queryset(self):
        queryset = EnterpriseDocument.objects.filter(
            enterprise=self.request.user
        )
        
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category=category)
        
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        period = self.request.query_params.get('period')
        if period:
            queryset = queryset.filter(period__icontains=period)
        
        return queryset.order_by('-uploaded_at')


class EnterpriseDocumentUploadView(generics.CreateAPIView):
    """
    POST /api/enterprise/documents/upload/
    Upload un nouveau document (PDF, Excel, etc.)
    
    Multipart form data:
    - file: Fichier à uploader
    - category: Catégorie du document
    - title: Titre
    - period: Période (ex: "Q3 2024")
    - period_start: Date début (optionnel)
    - period_end: Date fin (optionnel)
    """
    permission_classes = [IsEnterpriseUser]
    serializer_class = EnterpriseDocumentUploadSerializer
    parser_classes = [MultiPartParser, FormParser]
    
    def perform_create(self, serializer):
        requested_category = serializer.validated_data.get('category', 'other')
        file_obj = serializer.validated_data.get('file')
        title = serializer.validated_data.get('title', '')
        inferred_category = _infer_enterprise_upload_category(
            requested_category,
            getattr(file_obj, 'name', ''),
            title,
        )
        serializer.save(category=inferred_category)


class EnterpriseDocumentAnalyzeView(APIView):
    """
    POST /api/enterprise/documents/<id>/analyze/
    Analyse un document entreprise à la demande et stocke une synthèse structurée.
    """
    permission_classes = [IsEnterpriseUser]

    def post(self, request, document_id):
        try:
            document = EnterpriseDocument.objects.get(id=document_id, enterprise=request.user)
        except EnterpriseDocument.DoesNotExist:
            return Response({'error': 'Document introuvable.'}, status=status.HTTP_404_NOT_FOUND)

        document.status = 'processing'
        document.save(update_fields=['status'])

        analysis_summary = _build_enterprise_document_analysis(document)
        parser_result = analysis_summary.get('parser_result')
        if parser_result:
            analysis_summary['parser_result'] = {
                'transactions_count': len(parser_result.get('transactions', [])),
                'transactions': parser_result.get('transactions', [])[:200],
                'quality': parser_result.get('quality', {}),
                'teras_signals': parser_result.get('teras_signals', {}),
                'recommendations': parser_result.get('recommendations', []),
            }

        document.analysis_summary = analysis_summary
        document.processed_at = timezone.now()
        document.status = 'validated'
        document.validation_notes = "Analyse à la demande terminée."
        document.save(update_fields=['analysis_summary', 'processed_at', 'status', 'validation_notes'])

        return Response({
            'message': "Analyse documentaire prête. Vous pouvez maintenant l'appliquer au moteur TERAS.",
            'document': EnterpriseDocumentSerializer(document, context={'request': request}).data,
            'analysis_summary': analysis_summary,
        }, status=status.HTTP_200_OK)


class EnterpriseDocumentApplyView(APIView):
    """
    POST /api/enterprise/documents/<id>/apply/
    Applique les analyses documentaires au moteur TERAS Entreprise.
    """
    permission_classes = [IsEnterpriseUser]

    def post(self, request, document_id):
        try:
            document = EnterpriseDocument.objects.get(id=document_id, enterprise=request.user)
        except EnterpriseDocument.DoesNotExist:
            return Response({'error': 'Document introuvable.'}, status=status.HTTP_404_NOT_FOUND)

        if _enterprise_document_needs_refresh(document):
            return Response(
                {
                    'error': (
                        "Analyse requise avant application. Cliquez d'abord sur "
                        "'Analyser' pour générer l'interprétation du document."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        generated_score, dashboard_updates = _recompute_enterprise_from_documents(request.user, current_document=document)

        document.analysis_summary = {
            **(document.analysis_summary or {}),
            'applied_to_teras': True,
            'applied_at': timezone.now().isoformat(),
            'applied_score': {
                'id': generated_score.id,
                'score': generated_score.score,
                'breakdown': generated_score.breakdown,
                'computed_at': generated_score.computed_at.isoformat(),
            },
            'dashboard_updates': dashboard_updates,
        }
        document.processed_at = timezone.now()
        document.status = 'validated'
        document.validation_notes = "Analyse appliquée au moteur TERAS."
        document.save(update_fields=['analysis_summary', 'processed_at', 'status', 'validation_notes'])

        return Response({
            'message': (
                f"Analyse appliquée au moteur TERAS. Nouveau score entreprise : "
                f"{generated_score.score}/1000."
            ),
            'document': EnterpriseDocumentSerializer(document, context={'request': request}).data,
            'score': {
                'id': generated_score.id,
                'value': generated_score.score,
                'breakdown': generated_score.breakdown,
            },
            'dashboard_updates': dashboard_updates,
        }, status=status.HTTP_200_OK)


# Suite dans views_enterprise_part2.py...
