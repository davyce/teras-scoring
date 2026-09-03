# backend/scoring/views_documents.py
"""
TERAS Document Views — Endpoints upload, parsing, analyse
Intègre le pipeline document_parser.py complet.

Endpoints :
  POST /api/scoring/user/documents/upload/       → Upload + parsing auto
  GET  /api/scoring/user/documents/list/         → Liste documents
  GET  /api/scoring/user/documents/<id>/         → Détail + résultats parsing
  DELETE /api/scoring/user/documents/<id>/delete/ → Supprimer
  POST /api/scoring/user/documents/<id>/analyze/ → Analyser avec IA
  GET  /api/scoring/user/documents/<id>/download/ → Télécharger
  POST /api/scoring/user/documents/<id>/apply/   → Appliquer au score TERAS
"""

import os
import json
import threading
import logging
import re
import csv
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from django.conf import settings
from django.core.files.storage import default_storage
from django.http import FileResponse, JsonResponse
from django.db import models
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from .document_parser import parse_document, detect_format
from .models import Asset, Income, Recommendation, TerasScore, Transaction, UserDocument

logger = logging.getLogger(__name__)

# Taille max synchrone : 2MB → au-delà on parse en thread background
SYNC_MAX_BYTES = 2 * 1024 * 1024   # 2 MB
MAX_UPLOAD_BYTES = 15 * 1024 * 1024  # 15 MB

ALLOWED_EXTENSIONS = {
    'pdf', 'xlsx', 'xls', 'csv',
    'ofx', 'qfx', 'qif', 'sta', 'mt940',
    'jpg', 'jpeg', 'png', 'tiff',
}

ALLOWED_MIME_TYPES = {
    'application/pdf',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'application/vnd.ms-excel',
    'text/csv', 'text/plain',
    'image/jpeg', 'image/png', 'image/tiff',
    'application/ofx',
    'application/x-qif',
    'application/octet-stream',  # Pour OFX/QIF/MT940 sans MIME spécifique
}

PARSE_EXPECTED_DOC_TYPES = {
    'bank_statement',
    'invoice',
}


def _finalize_parse_status(result: dict, *, parse_expected: bool, uploaded_message: str) -> dict:
    """Normalise le statut après parsing pour distinguer stockage et échec réel."""
    errors = [str(err).strip() for err in result.get('errors', []) if str(err).strip()]
    parsing_success = bool(result.get('parsing_success'))
    analysis_ready = bool(result.get('analysis_summary'))

    if parsing_success or analysis_ready:
        status_value = 'parsed'
        display_status = 'analyzed'
        message = 'Document enregistré et analysé avec succès.'
    elif errors and parse_expected:
        status_value = 'failed'
        display_status = 'failed'
        message = 'Document enregistré, mais l’analyse structurée a échoué.'
    else:
        status_value = 'uploaded'
        display_status = 'stored'
        message = uploaded_message

    result['status'] = status_value
    result['display_status'] = display_status
    result['parse_expected'] = parse_expected
    result['message'] = message
    return result


def _safe_float(value, default: float = 0.0) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return default


def _doc_type_to_category(doc_type: str) -> str:
    return {
        'bank_statement': 'bank_statement',
        'salary_slip': 'payslip',
        'invoice': 'invoice',
        'proof_asset': 'proof_asset',
        'identity': 'identity',
        'tax_document': 'tax_document',
    }.get(doc_type or 'other', 'other')


def _infer_user_doc_type(requested_doc_type: str, filename: str, description: str = '') -> str:
    explicit = (requested_doc_type or 'other').strip().lower()
    raw = ' '.join(filter(None, [explicit, filename or '', description or ''])).lower()
    raw = raw.replace('-', '_').replace(' ', '_')

    if explicit in {'proof_asset', 'identity', 'tax_document', 'salary_slip', 'invoice'}:
        return explicit

    if any(token in raw for token in (
        'vehicle_registration',
        'carte_grise',
        'registration',
        'asset_declaration',
        'asset_inventory',
        'inventaire',
        'fixed_assets',
        'savings_proof',
        'proof_asset',
        'titre_de_propriete',
        'property_title',
        'lease_title',
        'collateral',
        'nantissement',
        'asset_statement',
    )):
        return 'proof_asset'

    if any(token in raw for token in (
        'identity',
        'id_card',
        'identity_card',
        'carte_identite',
        'passport',
        'passeport',
        'driver_license',
        'permis',
    )):
        return 'identity'

    if any(token in raw for token in (
        'tax',
        'fiscal',
        'impot',
        'patente',
        'cnss',
        'tax_document',
    )):
        return 'tax_document'

    if any(token in raw for token in (
        'salary',
        'bulletin',
        'paie',
        'payroll',
        'employment_certificate',
        'fiche_de_paie',
    )):
        return 'salary_slip'

    if any(token in raw for token in (
        'invoice',
        'facture',
        'receipt',
        'recu',
    )):
        return 'invoice'

    if any(token in raw for token in (
        'bank_statement',
        'income_expenses',
        'statement',
        'releve',
        'mt940',
        'ofx',
        'qif',
        'sales_register',
        'purchase_register',
    )):
        return 'bank_statement'

    return explicit or 'other'


def _relative_media_path(file_path: str) -> str:
    try:
        return os.path.relpath(file_path, settings.MEDIA_ROOT).replace(os.sep, '/')
    except Exception:
        return file_path


def _sync_user_document_record(user, file_path: str, doc_id: str, *, doc_type: str, parse_result: dict | None = None,
                               analysis_payload: dict | None = None, generated_score=None):
    parse_result = parse_result or {}
    analysis_payload = analysis_payload or {}
    mime_type = parse_result.get('mime_type') or ''
    errors = [str(err).strip() for err in parse_result.get('errors', []) if str(err).strip()]
    authenticity = _safe_float(parse_result.get('quality', {}).get('authenticity_score'))
    relative_path = _relative_media_path(file_path)

    document, _ = UserDocument.objects.get_or_create(
        user=user,
        filename=doc_id,
        defaults={
            'file': relative_path,
            'file_size': os.path.getsize(file_path) if os.path.exists(file_path) else 0,
            'mime_type': mime_type or 'application/octet-stream',
            'status': parse_result.get('status', 'uploaded'),
            'category': _doc_type_to_category(doc_type),
        },
    )

    updated_fields = []
    if not document.file:
        document.file = relative_path
        updated_fields.append('file')
    if document.file_size != (os.path.getsize(file_path) if os.path.exists(file_path) else document.file_size):
        document.file_size = os.path.getsize(file_path)
        updated_fields.append('file_size')

    new_status = parse_result.get('status') or document.status
    if document.status != new_status:
        document.status = new_status
        updated_fields.append('status')

    new_category = _doc_type_to_category(doc_type)
    if document.category != new_category:
        document.category = new_category
        updated_fields.append('category')

    if mime_type and document.mime_type != mime_type:
        document.mime_type = mime_type
        updated_fields.append('mime_type')

    if parse_result:
        document.extracted_data = parse_result
        document.confidence = max(document.confidence, min(1.0, authenticity))
        document.error_message = '\n'.join(errors)
        updated_fields.extend(['extracted_data', 'confidence', 'error_message'])

    if analysis_payload:
        document.ai_analysis = analysis_payload
        document.processed_at = timezone.now()
        updated_fields.extend(['ai_analysis', 'processed_at'])

    if generated_score is not None and document.generated_score_id != generated_score.id:
        document.generated_score = generated_score
        updated_fields.append('generated_score')

    if updated_fields:
        document.save(update_fields=list(dict.fromkeys(updated_fields)))

    return document


def _normalize_user_document_text(value) -> str:
    return ' '.join(str(value or '').replace('\xa0', ' ').split())


def _extract_fcfa_amounts_with_context(text: str) -> list[dict]:
    matches = []
    for match in re.finditer(r'(\d[\d\s.,]{2,})\s*(?:FCFA|XAF)', text or '', flags=re.IGNORECASE):
        raw_amount = match.group(1)
        cleaned = raw_amount.replace(' ', '').replace('\u202f', '')
        if ',' in cleaned and '.' in cleaned:
            cleaned = cleaned.replace('.', '').replace(',', '.')
        elif ',' in cleaned:
            cleaned = cleaned.replace(',', '.')
        try:
            amount = float(cleaned)
        except ValueError:
            continue
        if amount < 10_000:
            continue
        matches.append({
            'amount': amount,
            'start': match.start(),
            'end': match.end(),
            'raw': raw_amount,
        })
    return matches


def _guess_asset_type(text: str, filename: str = '') -> str:
    raw = f"{text or ''} {filename or ''}".lower()
    if any(token in raw for token in ('vehicule', 'vehicle', 'voiture', 'carte grise', 'registration', 'fourgonnette', 'camion')):
        return 'vehicule'
    if any(token in raw for token in ('epargne', 'savings', 'solde', 'depot', 'deposit', 'banque')):
        return 'epargne'
    if any(token in raw for token in ('immobilier', 'property', 'maison', 'appartement', 'titre de propriete', 'terrain', 'parcelle')):
        return 'immobilier' if 'terrain' not in raw and 'parcelle' not in raw else 'terrain'
    if any(token in raw for token in ('equipement', 'machine', 'materiel', 'ordinateur', 'stock', 'outil')):
        return 'equipement'
    return 'autre'


def _guess_asset_proof_mode(filename: str, text: str = '') -> str:
    raw = f"{filename or ''} {text or ''}".lower().replace('-', '_').replace(' ', '_')
    if 'vehicle_registration' in raw or 'carte_grise' in raw or 'registration' in raw:
        return 'vehicle_registration'
    if 'savings_proof' in raw or 'proof_epargne' in raw or 'solde' in raw:
        return 'savings_proof'
    if 'property_title' in raw or 'titre_de_propriete' in raw:
        return 'property_title'
    if 'asset_inventory' in raw or 'inventaire' in raw:
        return 'asset_inventory'
    if 'asset_declaration' in raw or 'declaration_d_actifs' in raw or 'declaration_actifs' in raw:
        return 'asset_declaration'
    return ''


def _default_asset_label(asset_type: str, proof_mode: str, filename: str) -> str:
    if proof_mode == 'vehicle_registration':
        return 'Carte grise ou titre véhicule'
    if proof_mode == 'savings_proof':
        return "Justificatif d'épargne"
    if proof_mode == 'property_title':
        return 'Titre de propriété'
    if proof_mode == 'asset_inventory':
        return "Inventaire d'actifs"
    if proof_mode == 'asset_declaration':
        return "Déclaration d'actifs"
    return {
        'vehicule': 'Actif véhicule documenté',
        'epargne': 'Épargne documentée',
        'immobilier': 'Bien immobilier documenté',
        'terrain': 'Terrain documenté',
        'equipement': 'Équipement documenté',
    }.get(asset_type, Path(filename or 'document').stem.replace('_', ' '))


def _read_user_document_text(file_path: str) -> str:
    suffix = Path(file_path).suffix.lower()
    try:
        if suffix == '.pdf':
            from pypdf import PdfReader

            reader = PdfReader(file_path)
            return '\n'.join((page.extract_text() or '') for page in reader.pages[:5])[:20000]
        if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
            from openpyxl import load_workbook

            workbook = load_workbook(file_path, data_only=True, read_only=True)
            rows = []
            for sheet in workbook.worksheets[:2]:
                for row in sheet.iter_rows(min_row=1, max_row=80, values_only=True):
                    rows.append(' | '.join(_normalize_user_document_text(cell) for cell in row if cell not in (None, '')))
            return '\n'.join(filter(None, rows))[:20000]
        if suffix == '.csv':
            return Path(file_path).read_text(encoding='utf-8', errors='ignore')[:20000]
        if suffix in {'.txt', '.md'}:
            return Path(file_path).read_text(encoding='utf-8', errors='ignore')[:20000]
    except Exception as exc:
        logger.warning("Lecture texte document utilisateur impossible pour %s: %s", file_path, exc)
    return ''


def _extract_assets_from_spreadsheet(file_path: str, filename: str) -> list[dict]:
    suffix = Path(file_path).suffix.lower()
    rows = []
    try:
        if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
            from openpyxl import load_workbook

            workbook = load_workbook(file_path, data_only=True, read_only=True)
            for sheet in workbook.worksheets[:2]:
                for row in sheet.iter_rows(min_row=1, max_row=100, values_only=True):
                    rows.append([_normalize_user_document_text(cell) for cell in row])
        elif suffix == '.csv':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as csv_file:
                sample = csv_file.read(2048)
                csv_file.seek(0)
                dialect = csv.Sniffer().sniff(sample, delimiters=';,')
                reader = csv.reader(csv_file, dialect)
                rows = [[_normalize_user_document_text(cell) for cell in row] for row in reader]
        else:
            return []
    except Exception as exc:
        logger.warning("Extraction tableur actif impossible pour %s: %s", file_path, exc)
        return []

    header_index = None
    label_idx = desc_idx = value_idx = proof_idx = None
    for row_index, row in enumerate(rows[:8]):
        normalized = [cell.lower() for cell in row]
        for index, cell in enumerate(normalized):
            if label_idx is None and any(token in cell for token in ('actif', 'asset', 'type')):
                label_idx = index
            if desc_idx is None and 'description' in cell:
                desc_idx = index
            if value_idx is None and any(token in cell for token in ('valeur', 'value', 'estime', 'estimated', 'montant', 'amount')):
                value_idx = index
            if proof_idx is None and any(token in cell for token in ('preuve', 'proof', 'mode')):
                proof_idx = index
        if value_idx is not None:
            header_index = row_index
            break

    if value_idx is None:
        return []

    items = []
    for row in rows[(header_index or 0) + 1:]:
        if value_idx >= len(row):
            continue
        raw_value = row[value_idx]
        try:
            value = float(str(raw_value).replace(' ', '').replace(',', '.'))
        except ValueError:
            continue
        if value < 10_000:
            continue
        label = row[label_idx] if label_idx is not None and label_idx < len(row) else ''
        description = row[desc_idx] if desc_idx is not None and desc_idx < len(row) else ''
        proof_mode = row[proof_idx] if proof_idx is not None and proof_idx < len(row) else ''
        asset_type = _guess_asset_type(f"{label} {description}", filename)
        proof_mode = _guess_asset_proof_mode(filename, proof_mode or description or label)
        items.append({
            'asset_type': asset_type,
            'label': (label or description or _default_asset_label(asset_type, proof_mode, filename))[:160],
            'description': (description or label or _default_asset_label(asset_type, proof_mode, filename))[:255],
            'estimated_value_xaf': round(value),
            'proof_mode': proof_mode,
        })
    return items


def _extract_structured_assets_from_text(text: str, filename: str) -> list[dict]:
    lines = [_normalize_user_document_text(line) for line in (text or '').splitlines()]
    lines = [line for line in lines if line]
    if not lines:
        return []

    doc_level_proof_mode = _guess_asset_proof_mode(filename)
    items = []
    for index, line in enumerate(lines):
        amount_matches = _extract_fcfa_amounts_with_context(line)
        if not amount_matches:
            continue

        amount = round(amount_matches[0]['amount'])
        if amount < 10_000:
            continue

        asset_type_line = lines[index - 2] if index >= 2 else ''
        description_line = lines[index - 1] if index >= 1 else asset_type_line
        proof_hint = lines[index + 1] if index + 1 < len(lines) else ''
        secondary_hint = lines[index + 2] if index + 2 < len(lines) else ''
        context = ' '.join(filter(None, [asset_type_line, description_line, proof_hint, secondary_hint]))
        proof_mode = _guess_asset_proof_mode(filename, f'{proof_hint} {secondary_hint}')
        if not proof_mode and doc_level_proof_mode in {'asset_declaration', 'asset_inventory'}:
            proof_mode = doc_level_proof_mode
        asset_type = _guess_asset_type(context, filename)
        label = asset_type_line or _default_asset_label(asset_type, proof_mode, filename)
        description = description_line or label

        items.append({
            'asset_type': asset_type,
            'label': label[:160],
            'description': description[:255],
            'estimated_value_xaf': amount,
            'proof_mode': proof_mode,
        })

    return items


def _extract_user_asset_items(file_path: str, filename: str, doc_type: str) -> list[dict]:
    if doc_type != 'proof_asset':
        return []

    suffix = Path(file_path).suffix.lower()
    if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm', '.csv'}:
        spreadsheet_items = _extract_assets_from_spreadsheet(file_path, filename)
        if spreadsheet_items:
            return spreadsheet_items

    text = _read_user_document_text(file_path)
    if not text:
        return []

    structured_items = _extract_structured_assets_from_text(text, filename)
    if structured_items:
        return structured_items

    amount_matches = _extract_fcfa_amounts_with_context(text)
    if not amount_matches:
        return []

    filename_key = filename.lower().replace('-', '_')
    multi_asset_doc = any(token in filename_key for token in ('asset_declaration', 'asset_inventory', 'inventaire'))
    doc_level_proof_mode = _guess_asset_proof_mode(filename)
    items = []

    for index, match in enumerate(amount_matches[:6]):
        if not multi_asset_doc and index > 0:
            break
        context = text[max(0, match['start'] - 140):min(len(text), match['end'] + 80)]
        asset_type = _guess_asset_type(context, filename)
        proof_mode = _guess_asset_proof_mode(filename, context)
        if multi_asset_doc and not proof_mode and doc_level_proof_mode:
            proof_mode = doc_level_proof_mode
        description_match = re.search(
            r'Description\s+(.{5,140}?)(?:\s+Valeur estim|(?:\s+\d[\d\s.,]{2,}\s*(?:FCFA|XAF)))',
            context,
            flags=re.IGNORECASE | re.DOTALL,
        )
        label = _default_asset_label(asset_type, proof_mode, filename)
        description = _normalize_user_document_text(description_match.group(1)) if description_match else label
        items.append({
            'asset_type': asset_type,
            'label': label[:160],
            'description': description[:255],
            'estimated_value_xaf': round(match['amount']),
            'proof_mode': proof_mode,
        })

    deduped = []
    seen = set()
    for item in items:
        signature = (
            item.get('proof_mode') or '',
            item.get('asset_type') or '',
            int(round(_safe_float(item.get('estimated_value_xaf')))),
        )
        if signature in seen:
            continue
        seen.add(signature)
        deduped.append(item)
    return deduped


def _normalize_asset_record_key(item: dict, index: int) -> str:
    proof_mode = str(item.get('proof_mode') or '').strip().lower()
    asset_type = str(item.get('asset_type') or 'autre').strip().lower()
    amount = int(round(_safe_float(item.get('estimated_value_xaf'))))
    if proof_mode and proof_mode not in {'asset_declaration', 'asset_inventory'}:
        base = proof_mode
    else:
        base = f"{asset_type}_{amount}_{index + 1}"
    return re.sub(r'[^a-z0-9_]+', '_', base).strip('_') or f"asset_{index + 1}"


def _sync_user_assets_from_analysis(user, analysis_payload: dict, doc_id: str) -> dict:
    asset_items = list(analysis_payload.get('detected_assets') or [])
    if not asset_items:
        fallback_value = _safe_float(
            analysis_payload.get('summary', {}).get('estimated_asset_value_xaf')
            or analysis_payload.get('dashboard_updates', {}).get('estimated_asset_value_xaf')
        )
        if fallback_value > 0:
            asset_items = [{
                'asset_type': 'autre',
                'label': _infer_user_doc_type('proof_asset', doc_id),
                'description': _default_asset_label('autre', _guess_asset_proof_mode(doc_id), doc_id),
                'estimated_value_xaf': round(fallback_value),
                'proof_mode': _guess_asset_proof_mode(doc_id),
            }]

    synced_values = []
    for index, item in enumerate(asset_items):
        value = _safe_float(item.get('estimated_value_xaf'))
        if value <= 0:
            continue
        asset_type = str(item.get('asset_type') or 'autre').strip().lower()
        if asset_type not in {'immobilier', 'vehicule', 'terrain', 'equipement', 'epargne', 'autre'}:
            asset_type = 'autre'
        record_key = _normalize_asset_record_key(item, index)
        label = _default_asset_label(asset_type, str(item.get('proof_mode') or ''), doc_id)
        description = f"[ASSET:{record_key}] {label}"
        Asset.objects.filter(
            user=user,
            asset_type=asset_type,
            estimated_value=Decimal(str(round(value, 2))),
            verified=True,
            description__startswith='[ASSET:',
        ).exclude(description=description).delete()
        Asset.objects.update_or_create(
            user=user,
            description=description,
            defaults={
                'asset_type': asset_type,
                'estimated_value': Decimal(str(round(value, 2))),
                'verified': True,
            },
        )
        synced_values.append(value)

    return {
        'assets_count': len(synced_values),
        'assets_total_xaf': round(sum(synced_values)),
    }


def _prepare_user_document_result(file_path: str, filename: str, result: dict) -> tuple[dict, bool]:
    normalized = dict(result or {})
    inferred_doc_type = _infer_user_doc_type(
        normalized.get('doc_type', 'other'),
        filename,
        normalized.get('description', ''),
    )
    changed = False
    if normalized.get('doc_type') != inferred_doc_type:
        normalized['doc_type'] = inferred_doc_type
        changed = True

    parse_expected = inferred_doc_type in PARSE_EXPECTED_DOC_TYPES
    if normalized.get('status') != 'processing':
        finalized = _finalize_parse_status(
            normalized,
            parse_expected=parse_expected,
            uploaded_message='Document enregistré. Analyse structurée non applicable ou aucune donnée exploitable détectée.',
        )
        if finalized != normalized:
            changed = True
        normalized = finalized
    return normalized, changed


def _update_user_document_category(user_id: int, doc_id: str, doc_type: str):
    UserDocument.objects.filter(user_id=user_id, filename=doc_id).update(category=_doc_type_to_category(doc_type))


def _generate_local_user_analysis(user, parse_result: dict, *, file_path: str | None = None, doc_id: str | None = None) -> dict:
    doc_type = parse_result.get('doc_type', 'bank_statement')
    signals = parse_result.get('teras_signals', {}) or {}
    quality = parse_result.get('quality', {}) or {}
    stats = quality.get('stats', {}) or {}
    income_signal = signals.get('income_signal', {}) or {}
    savings_signal = signals.get('savings_signal', {}) or {}
    transactions_signal = signals.get('transactions_signal', {}) or {}

    avg_income = _safe_float(income_signal.get('monthly_avg_xaf'))
    crm_estimated = _safe_float(signals.get('crm_estimated_xaf'))
    savings_avg = _safe_float(savings_signal.get('monthly_deposit_avg_xaf'))
    months_covered = int(stats.get('months_covered') or signals.get('months_analyzed') or 0)
    authenticity = _safe_float(quality.get('authenticity_score'))
    regularity = _safe_float(transactions_signal.get('regularity_score'))
    stability = _safe_float(income_signal.get('income_stability'))
    streak = int(savings_signal.get('streak_months') or 0)
    net_cashflow = _safe_float(stats.get('net_cashflow_xaf'))

    strengths = []
    if avg_income > 0:
        strengths.append(f"revenu mensuel documente autour de {round(avg_income):,} FCFA")
    if regularity >= 0.65:
        strengths.append("flux de transactions relativement reguliers")
    if streak >= 3:
        strengths.append(f"effort d'epargne detecte sur {streak} mois")
    if authenticity >= 0.8:
        strengths.append("bon niveau d'authenticite documentaire")

    risks = []
    if months_covered < 3:
        risks.append("historique trop court pour consolider le profil")
    if net_cashflow < 0:
        risks.append("cashflow net negatif sur la periode analysee")
    if stability < 0.5 and avg_income > 0:
        risks.append("revenu peu stable d'un mois a l'autre")
    if authenticity < 0.55:
        risks.append("qualite documentaire a verifier manuellement")

    recommended_actions = [
        "maintenir des entrees regulieres sur le compte principal",
        "documenter davantage les revenus recurrents avec bulletins ou justificatifs",
        "conserver les preuves d'actifs pour renforcer le pilier A",
    ]
    if savings_avg > 0:
        recommended_actions[0] = "maintenir les depots reguliers pour consolider la capacite de remboursement"

    asset_items = []
    if doc_type == 'proof_asset' and file_path:
        asset_items = _extract_user_asset_items(file_path, doc_id or Path(file_path).name, doc_type)

    estimated_asset_value = 0
    if doc_type == 'proof_asset':
        extracted_asset_total = round(sum(_safe_float(item.get('estimated_value_xaf')) for item in asset_items))
        estimated_asset_value = max(
            extracted_asset_total,
            round(avg_income * 6),
            round(savings_avg * 10),
            round(crm_estimated * 8),
            350000,
        )
        strengths = [
            "preuve d'actif disponible pour renforcer le pilier Actifs",
            f"valeur d'actif exploitable estimee autour de {estimated_asset_value:,} FCFA",
        ]
        if asset_items:
            strengths.append(f"{len(asset_items)} actif(s) distinct(s) ont ete detectes dans la piece")
        if authenticity >= 0.7:
            strengths.append("coherence documentaire suffisante pour enrichir le passeport financier")
        risks = [
            "verification complementaire recommandee avant nantissement ou prise de garantie"
        ]
        recommended_actions = [
            "joindre la facture, la carte grise ou le titre de propriete associe pour renforcer la preuve",
            "indiquer si l'actif est libre de toute charge ou deja finance",
            "mettre a jour periodiquement la valeur estimee de l'actif dans le dossier TERAS",
        ]
    elif doc_type == 'identity':
        strengths = [
            "piece d'identite disponible pour fiabiliser l'identification du dossier",
            "le dossier KYC devient plus exploitable pour une instruction bancaire rapide",
        ]
        risks = ["verifier que la piece est encore valide et lisible"] if authenticity < 0.75 else []
        recommended_actions = [
            "completer avec un justificatif de domicile recent pour consolider le KYC",
            "verifier la concordance entre nom, NIU et numero de telephone declares",
        ]
    elif doc_type == 'tax_document':
        strengths = [
            "document fiscal disponible pour documenter les revenus ou l'activite declaree",
            "la transparence du dossier s'en trouve renforcee pour les contreparties bancaires",
        ]
        risks = ["periode fiscale ou reference administrative a confirmer"] if months_covered == 0 else []
        recommended_actions = [
            "ajouter la periode fiscale couverte et les references d'administration si elles figurent sur le document",
            "croiser ce document avec un releve bancaire ou une preuve de revenu pour un impact TERAS plus fort",
        ]
    elif doc_type == 'salary_slip' and avg_income <= 0:
        strengths = [
            "justificatif de revenu disponible meme sans extraction transactionnelle complete",
            "la regularite salariale peut etre consolidee avec plusieurs bulletins consecutifs",
        ]
        risks = ["montant net mensuel a confirmer avec d'autres bulletins ou credits bancaires"]
        recommended_actions = [
            "ajouter au moins trois bulletins consecutifs pour mieux objectiver la stabilite du revenu",
            "associer un releve bancaire recent pour faire converger les signaux revenus et transactions",
        ]

    estimated_change = int(
        min(
            140,
            max(
                18,
                (regularity * 35)
                + (stability * 30)
                + min(streak, 6) * 6
                + min(months_covered, 6) * 4
                + (authenticity * 25),
            ),
        )
    )
    pillars_affected = ['T', 'E', 'R']
    if savings_avg > 0:
        pillars_affected.append('A')
    if doc_type == 'proof_asset':
        pillars_affected = ['A', 'T']
        estimated_change = max(estimated_change, 42)
    elif doc_type == 'identity':
        pillars_affected = ['S', 'T']
        estimated_change = max(estimated_change, 16)
    elif doc_type == 'tax_document':
        pillars_affected = ['T', 'R']
        estimated_change = max(estimated_change, 22)
    elif doc_type == 'salary_slip' and avg_income <= 0:
        pillars_affected = ['R', 'E']
        estimated_change = max(estimated_change, 24)

    narrative = (
        "## Synthese\n"
        f"Le dossier de {user.get_full_name() or user.email} a ete relu a partir d'un document de type {doc_type}. "
        f"Le revenu mensuel exploitable ressort autour de {round(avg_income):,} FCFA "
        f"et une capacite de remboursement proche de {round(crm_estimated):,} FCFA/mois. "
        f"L'historique exploitable couvre {months_covered} mois avec un score d'authenticite de {round(authenticity * 100)}%.\n\n"
        "## Points forts\n"
        + ('\n'.join(f"- {item}" for item in (strengths or ['aucun point fort significatif automatiquement detecte'])) if True else '')
        + "\n\n## Points de vigilance\n"
        + ('\n'.join(f"- {item}" for item in (risks or ['aucun risque majeur automatiquement detecte'])) if True else '')
        + "\n\n## Actions recommandees\n"
        + '\n'.join(f"- {item}" for item in recommended_actions[:3])
    )

    return {
        'analysis_origin': 'local-heuristic',
        'document_type': doc_type,
        'analysis_text': narrative,
        'summary': {
            'monthly_income_xaf': round(avg_income),
            'monthly_savings_xaf': round(savings_avg),
            'crm_estimated_xaf': round(crm_estimated),
            'months_covered': months_covered,
            'net_cashflow_xaf': round(net_cashflow),
            'authenticity_score': authenticity,
            'estimated_asset_value_xaf': estimated_asset_value,
            'asset_items_count': len(asset_items),
        },
        'strengths': strengths,
        'risks': risks,
        'recommended_actions': recommended_actions,
        'detected_assets': asset_items,
        'score_impact': {
            'estimated_change': estimated_change,
            'pillars_affected': list(dict.fromkeys(pillars_affected)),
            'confidence': round(min(0.96, max(0.35, authenticity * 0.6 + regularity * 0.4)), 2),
        },
        'dashboard_updates': {
            'monthly_income_xaf': round(avg_income),
            'monthly_savings_xaf': round(savings_avg),
            'crm_estimated_xaf': round(crm_estimated),
            'transactions_count': int(stats.get('total_transactions') or len(parse_result.get('transactions', []))),
            'estimated_asset_value_xaf': estimated_asset_value,
            'asset_items_count': len(asset_items),
            'document_type': doc_type,
            'authenticity_score': round(authenticity, 3),
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# MODÈLE LÉGER (dict en mémoire / simple JSON dans le User)
# On utilise un modèle Django simplifié via JSONField si disponible
# ou on stocke les résultats dans un fichier .json à côté du document
# ─────────────────────────────────────────────────────────────────────────────

def _get_result_path(doc_path: str) -> str:
    """Chemin du fichier JSON de résultats (à côté du document)."""
    return doc_path + '.result.json'


def _save_result(doc_path: str, result: dict):
    """Sauvegarde le résultat de parsing en JSON."""
    try:
        result_path = _get_result_path(doc_path)
        with open(result_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2, default=str)
    except Exception as e:
        logger.error(f"Erreur sauvegarde résultat : {e}")


def _load_result(doc_path: str) -> dict:
    """Charge le résultat de parsing depuis le JSON."""
    result_path = _get_result_path(doc_path)
    try:
        if os.path.exists(result_path):
            with open(result_path, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def _get_user_doc_dir(user_id: int) -> str:
    """Retourne (et crée) le répertoire de documents de l'utilisateur."""
    doc_dir = os.path.join(
        getattr(settings, 'MEDIA_ROOT', 'media'),
        'documents',
        str(user_id)
    )
    os.makedirs(doc_dir, exist_ok=True)
    return doc_dir


def _list_user_docs(user_id: int) -> list:
    """Liste tous les documents d'un utilisateur."""
    doc_dir = _get_user_doc_dir(user_id)
    docs    = []

    try:
        for fname in os.listdir(doc_dir):
            # Ignorer les fichiers de résultats
            if fname.endswith('.result.json'):
                continue

            ext = Path(fname).suffix.lower().lstrip('.')
            if ext not in ALLOWED_EXTENSIONS:
                continue

            fpath  = os.path.join(doc_dir, fname)
            stat   = os.stat(fpath)
            result, changed = _prepare_user_document_result(fpath, fname, _load_result(fpath))
            doc_type = result.get('doc_type', 'other')
            if changed:
                _save_result(fpath, result)
                _update_user_document_category(user_id, fname, doc_type)

            docs.append({
                'id':              fname,  # Le nom du fichier sert d'ID
                'filename':        fname,
                'size_bytes':      stat.st_size,
                'size_mb':         round(stat.st_size / (1024 * 1024), 2),
                'format':          result.get('format', detect_format(fpath, fname)),
                'uploaded_at':     datetime.fromtimestamp(stat.st_ctime).isoformat(),
                'parsing_success': result.get('parsing_success', False),
                'status':          result.get('status', 'uploaded'),
                'transactions_count': len(result.get('transactions', [])),
                'authenticity_score': result.get('quality', {}).get('authenticity_score', 0),
                'crm_estimated':      result.get('teras_signals', {}).get('crm_estimated_xaf', 0),
                'months_covered':     result.get('quality', {}).get('stats', {}).get('months_covered', 0),
                'errors_count':       len(result.get('errors', [])),
                'doc_type':           doc_type,
                'parse_expected':     result.get(
                    'parse_expected',
                    result.get('doc_type', 'bank_statement') in PARSE_EXPECTED_DOC_TYPES,
                ),
                'display_status':     result.get('display_status', result.get('status', 'uploaded')),
                'message':            result.get('message', ''),
            })
    except Exception as e:
        logger.error(f"Erreur listage documents user {user_id} : {e}")

    # Trier par date (plus récent en premier)
    docs.sort(key=lambda d: d['uploaded_at'], reverse=True)
    return docs


# ─────────────────────────────────────────────────────────────────────────────
# ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
@permission_classes([IsAuthenticated])
def upload_document(request):
    """
    POST /api/scoring/user/documents/upload/

    Upload d'un document financier + parsing automatique.

    Body (multipart) :
      file         : Le fichier
      doc_type     : 'bank_statement' | 'salary_slip' | 'invoice' | 'other'
      description  : (optionnel) Note utilisateur
      apply_to_score : '1' pour appliquer immédiatement au score TERAS

    Réponse :
      200 : {status, filename, format, transactions_count, teras_signals, quality, recommendations}
      400 : {error}
    """
    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'Aucun fichier fourni.'}, status=400)

    # ── Validation ──────────────────────────────────────────────────
    filename = file.name
    ext      = Path(filename).suffix.lower().lstrip('.')

    if ext not in ALLOWED_EXTENSIONS:
        return Response({
            'error': f"Extension '{ext}' non supportée.",
            'allowed': list(ALLOWED_EXTENSIONS),
        }, status=400)

    if file.size > MAX_UPLOAD_BYTES:
        return Response({
            'error': f"Fichier trop volumineux ({round(file.size / 1024 / 1024, 1)} MB). Maximum : 15 MB.",
        }, status=400)

    # ── Sauvegarde ──────────────────────────────────────────────────
    doc_dir = _get_user_doc_dir(request.user.id)

    # Nom unique avec timestamp
    ts        = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name = f"{ts}_{filename.replace(' ', '_')}"
    file_path = os.path.join(doc_dir, safe_name)

    try:
        with open(file_path, 'wb') as f:
            for chunk in file.chunks():
                f.write(chunk)
    except Exception as e:
        return Response({'error': f"Erreur sauvegarde : {e}"}, status=500)

    # ── Parsing ─────────────────────────────────────────────────────
    doc_type    = request.data.get('doc_type', 'bank_statement')
    doc_type    = _infer_user_doc_type(doc_type, filename, request.data.get('description', ''))
    description = request.data.get('description', '')
    apply_score = request.data.get('apply_to_score', '0') == '1'
    mime_type   = file.content_type or ''
    parse_expected = doc_type in PARSE_EXPECTED_DOC_TYPES

    # Sync < 2MB, async sinon
    if file.size <= SYNC_MAX_BYTES:
        # Parsing synchrone
        try:
            result = parse_document(file_path, filename, mime_type)
        except Exception as exc:
            result = {
                'errors': [str(exc)],
                'parsing_success': False,
                'transactions': [],
                'quality': {},
                'teras_signals': {},
                'recommendations': [],
                'format': detect_format(file_path, filename, mime_type),
            }
        result['doc_type']    = doc_type
        result['description'] = description
        result = _finalize_parse_status(
            result,
            parse_expected=parse_expected,
            uploaded_message='Document enregistré. Analyse structurée non applicable ou aucune donnée exploitable détectée.',
        )
        _save_result(file_path, result)
        _sync_user_document_record(
            request.user,
            file_path,
            safe_name,
            doc_type=doc_type,
            parse_result=result,
        )

        # Appliquer au score si demandé
        if apply_score and result['parsing_success']:
            apply_msg = _apply_to_score(request.user, result, doc_id=safe_name, file_path=file_path)
        else:
            apply_msg = None

        return Response({
            'status':             result['status'],
            'filename':           safe_name,
            'format':             result['format'],
            'transactions_count': len(result['transactions']),
            'authenticity_score': result['quality'].get('authenticity_score', 0),
            'teras_signals':      result['teras_signals'],
            'quality_stats':      result['quality'].get('stats', {}),
            'recommendations':    result['recommendations'],
            'errors':             result['errors'],
            'applied_to_score':   apply_msg,
            'parsing_mode':       'synchronous',
            'message':            result['message'],
            'parse_expected':     result['parse_expected'],
            'display_status':     result['display_status'],
        })

    else:
        # Parsing asynchrone (fichier > 2MB)
        initial_result = {
            'status':      'processing',
            'display_status': 'processing',
            'format':      detect_format(file_path, filename, mime_type),
            'doc_type':    doc_type,
            'description': description,
            'filename':    safe_name,
            'parse_expected': parse_expected,
            'message':     'Fichier volumineux. Analyse en arrière-plan en cours.',
        }
        _save_result(file_path, initial_result)
        _sync_user_document_record(
            request.user,
            file_path,
            safe_name,
            doc_type=doc_type,
            parse_result=initial_result,
        )

        def _parse_async():
            try:
                result = parse_document(file_path, filename, mime_type)
                result['doc_type']    = doc_type
                result['description'] = description
                result = _finalize_parse_status(
                    result,
                    parse_expected=parse_expected,
                    uploaded_message='Document enregistré. Analyse structurée non applicable ou aucune donnée exploitable détectée.',
                )
                _save_result(file_path, result)
                _sync_user_document_record(
                    request.user,
                    file_path,
                    safe_name,
                    doc_type=doc_type,
                    parse_result=result,
                )
                logger.info(f"[TERAS] Async parsing terminé : {safe_name}")
            except Exception as e:
                err_result = {
                    'status': 'failed',
                    'display_status': 'failed',
                    'errors': [str(e)],
                    'doc_type': doc_type,
                    'description': description,
                    'parse_expected': parse_expected,
                    'message': 'Document enregistré, mais l’analyse structurée a échoué.',
                }
                _save_result(file_path, err_result)
                _sync_user_document_record(
                    request.user,
                    file_path,
                    safe_name,
                    doc_type=doc_type,
                    parse_result=err_result,
                )
                logger.error(f"[TERAS] Erreur async parsing {safe_name} : {e}")

        thread = threading.Thread(target=_parse_async, daemon=True)
        thread.start()

        return Response({
            'status':       'processing',
            'filename':     safe_name,
            'format':       initial_result['format'],
            'message':      'Fichier volumineux. Parsing en cours en arrière-plan. Revenez dans quelques secondes.',
            'polling_url':  f'/api/scoring/user/documents/{safe_name}/',
            'parsing_mode': 'asynchronous',
            'parse_expected': parse_expected,
            'display_status': 'processing',
        })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_documents(request):
    """
    GET /api/scoring/user/documents/list/

    Liste tous les documents de l'utilisateur.
    """
    docs = _list_user_docs(request.user.id)
    return Response({
        'documents': docs,
        'count':     len(docs),
        'summary': {
            'total_parsed':    sum(1 for d in docs if d['status'] == 'parsed'),
            'total_failed':    sum(1 for d in docs if d['status'] == 'failed'),
            'total_processing':sum(1 for d in docs if d['status'] == 'processing'),
            'total_uploaded':  sum(1 for d in docs if d['status'] == 'uploaded'),
        }
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def document_detail(request, doc_id):
    """
    GET /api/scoring/user/documents/<doc_id>/

    Détail d'un document + résultats parsing complets.
    Utilisé pour le polling (status: processing → parsed).
    """
    doc_dir   = _get_user_doc_dir(request.user.id)
    file_path = os.path.join(doc_dir, doc_id)

    if not os.path.exists(file_path):
        return Response({'error': 'Document non trouvé.'}, status=404)

    result, changed = _prepare_user_document_result(file_path, doc_id, _load_result(file_path))
    doc_type = result.get('doc_type', 'other')
    if changed:
        _save_result(file_path, result)
        _sync_user_document_record(request.user, file_path, doc_id, doc_type=doc_type, parse_result=result)
    stat   = os.stat(file_path)

    # Réponse allégée (sans les transactions brutes pour la liste)
    include_txns = request.GET.get('include_transactions', '0') == '1'

    response_data = {
        'id':                 doc_id,
        'filename':           doc_id,
        'size_bytes':         stat.st_size,
        'format':             result.get('format', 'unknown'),
        'status':             result.get('status', 'uploaded'),
        'display_status':     result.get('display_status', result.get('status', 'uploaded')),
        'parse_expected':     result.get(
            'parse_expected',
            result.get('doc_type', 'bank_statement') in PARSE_EXPECTED_DOC_TYPES,
        ),
        'doc_type':           doc_type,
        'description':        result.get('description', ''),
        'uploaded_at':        datetime.fromtimestamp(stat.st_ctime).isoformat(),
        'parsed_at':          result.get('parsed_at'),
        'parsing_success':    result.get('parsing_success', False),
        'transactions_count': len(result.get('transactions', [])),
        'quality':            result.get('quality', {}),
        'teras_signals':      result.get('teras_signals', {}),
        'recommendations':    result.get('recommendations', []),
        'metadata':           result.get('metadata', {}),
        'errors':             result.get('errors', []),
        'message':            result.get('message', ''),
        'analysis_summary':   result.get('analysis_summary', {}),
        'analysis_text':      result.get('analysis_text', ''),
        'generated_score':    result.get('generated_score'),
        'confidence':         result.get('confidence', 0),
        'processed_at':       result.get('processed_at'),
    }

    if include_txns:
        response_data['transactions'] = result.get('transactions', [])[:100]  # Max 100

    return Response(response_data)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_document(request, doc_id):
    """
    DELETE /api/scoring/user/documents/<doc_id>/delete/
    """
    doc_dir   = _get_user_doc_dir(request.user.id)
    file_path = os.path.join(doc_dir, doc_id)

    if not os.path.exists(file_path):
        return Response({'error': 'Document non trouvé.'}, status=404)

    try:
        os.remove(file_path)
        result_path = _get_result_path(file_path)
        if os.path.exists(result_path):
            os.remove(result_path)
        UserDocument.objects.filter(user=request.user, filename=doc_id).delete()
        return Response({'message': 'Document supprimé.'})
    except Exception as e:
        return Response({'error': f"Erreur suppression : {e}"}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_document(request, doc_id):
    """
    GET /api/scoring/user/documents/<doc_id>/download/
    """
    doc_dir   = _get_user_doc_dir(request.user.id)
    file_path = os.path.join(doc_dir, doc_id)

    if not os.path.exists(file_path):
        return Response({'error': 'Document non trouvé.'}, status=404)

    try:
        return FileResponse(
            open(file_path, 'rb'),
            as_attachment=True,
            filename=doc_id,
        )
    except Exception as e:
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def analyze_document(request, doc_id):
    """
    POST /api/scoring/user/documents/<doc_id>/analyze/

    Analyse le document avec l'IA Claude Sonnet 4.
    Génère un commentaire sur les transactions et des conseils TERAS.
    """
    import requests as req

    doc_dir   = _get_user_doc_dir(request.user.id)
    file_path = os.path.join(doc_dir, doc_id)

    if not os.path.exists(file_path):
        return Response({'error': 'Document non trouvé.'}, status=404)

    result, changed = _prepare_user_document_result(file_path, doc_id, _load_result(file_path))
    if changed:
        _save_result(file_path, result)

    signals = result.get('teras_signals', {})
    quality = result.get('quality', {})
    stats   = quality.get('stats', {})
    user    = request.user

    prompt = f"""Tu es l'assistant IA TERAS, expert en crédit pour l'Afrique Centrale.
Analyse ce relevé bancaire de {user.get_full_name() or user.email} et fournis des conseils personnalisés.

RÉSULTATS DU PARSING :
- Transactions analysées : {stats.get('total_transactions', 0)}
- Période couverte : {stats.get('months_covered', 0)} mois ({stats.get('date_from')} → {stats.get('date_to')})
- Score d'authenticité : {quality.get('authenticity_score', 0) * 100:.0f}%
- Total entrées (crédits) : {stats.get('total_credits_xaf', 0):,.0f} FCFA
- Total sorties (débits)  : {stats.get('total_debits_xaf', 0):,.0f} FCFA
- Cashflow net : {stats.get('net_cashflow_xaf', 0):,.0f} FCFA
- Revenu mensuel moyen estimé : {signals.get('income_signal', {}).get('monthly_avg_xaf', 0):,.0f} FCFA
- CRM estimé (capacité remboursement) : {signals.get('crm_estimated_xaf', 0):,.0f} FCFA/mois
- Régularité transactions : {signals.get('transactions_signal', {}).get('regularity_score', 0) * 100:.0f}%
- Transactions tontine : {signals.get('tontine_transactions', 0)}
- Streak épargne : {signals.get('savings_signal', {}).get('streak_months', 0)} mois

INSTRUCTIONS :
1. Commence par un commentaire général sur la santé financière (2-3 phrases)
2. Identifie 2-3 points FORTS du profil (ce qui améliore le score TERAS)
3. Identifie 2-3 points à AMÉLIORER (avec impact chiffré en points TERAS)
4. Calcule et présente l'estimation du crédit maximum accessible (basé sur CRM)
5. Donne 3 conseils concrets pour améliorer le score dans les 3 prochains mois
6. Utilise des exemples locaux congolais. Sois chaleureux et encourageant.
7. Formatage : utilise ## pour les sections, - pour les listes
"""

    local_analysis = _generate_local_user_analysis(user, result, file_path=file_path, doc_id=doc_id)
    analysis_text = local_analysis['analysis_text']
    analysis_origin = local_analysis['analysis_origin']

    anthropic_key = getattr(settings, 'ANTHROPIC_API_KEY', '')
    if anthropic_key and result.get('parsing_success'):
        try:
            response = req.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key":         anthropic_key,
                    "content-type":      "application/json",
                    "anthropic-version": "2023-06-01",
                },
                json={
                    "model":      "claude-sonnet-4-20250514",
                    "max_tokens": 1500,
                    "messages":   [{"role": "user", "content": prompt}],
                },
                timeout=30,
            )
            response.raise_for_status()
            data = response.json()
            remote_text = data['content'][0]['text']
            if remote_text:
                analysis_text = remote_text
                analysis_origin = 'anthropic'
        except Exception as exc:
            logger.warning("Analyse Anthropic indisponible pour %s: %s", doc_id, exc)

    analysis_payload = {
        **local_analysis,
        'analysis_text': analysis_text,
        'analysis_origin': analysis_origin,
        'analyzed_at': timezone.now().isoformat(),
    }
    result['analysis_summary'] = analysis_payload
    result['analysis_text'] = analysis_text
    result['confidence'] = analysis_payload['score_impact']['confidence']
    result['processed_at'] = timezone.now().isoformat()
    result = _finalize_parse_status(
        result,
        parse_expected=result.get('doc_type', 'other') in PARSE_EXPECTED_DOC_TYPES,
        uploaded_message='Document enregistré. Analyse structurée non applicable ou aucune donnée exploitable détectée.',
    )
    _save_result(file_path, result)

    document = _sync_user_document_record(
        user,
        file_path,
        doc_id,
        doc_type=result.get('doc_type', 'bank_statement'),
        parse_result=result,
        analysis_payload=analysis_payload,
    )

    return Response({
        'analysis': analysis_text,
        'analysis_summary': analysis_payload,
        'teras_signals': signals,
        'quality': quality,
        'document_status': document.status,
        'analyzed_at': analysis_payload['analyzed_at'],
        'message': 'Analyse documentaire prête. Vous pouvez maintenant l’appliquer au moteur TERAS.',
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def apply_to_score(request, doc_id):
    """
    POST /api/scoring/user/documents/<doc_id>/apply/

    Applique les signaux parsés au score TERAS de l'utilisateur.
    Met à jour les piliers T, E, R basés sur les transactions du document.
    """
    doc_dir   = _get_user_doc_dir(request.user.id)
    file_path = os.path.join(doc_dir, doc_id)

    if not os.path.exists(file_path):
        return Response({'error': 'Document non trouvé.'}, status=404)

    result, changed = _prepare_user_document_result(file_path, doc_id, _load_result(file_path))
    if changed:
        _save_result(file_path, result)
    if not result.get('parsing_success') and not result.get('analysis_summary'):
        doc_type = result.get('doc_type', 'other')
        if doc_type == 'proof_asset':
            analysis_payload = _generate_local_user_analysis(request.user, result, file_path=file_path, doc_id=doc_id)
            result['analysis_summary'] = analysis_payload
            result['analysis_text'] = analysis_payload.get('analysis_text', '')
            result['processed_at'] = timezone.now().isoformat()
            result = _finalize_parse_status(
                result,
                parse_expected=False,
                uploaded_message='Document enregistré. Analyse structurée non applicable ou aucune donnée exploitable détectée.',
            )
            _save_result(file_path, result)
            _sync_user_document_record(
                request.user,
                file_path,
                doc_id,
                doc_type=doc_type,
                parse_result=result,
                analysis_payload=analysis_payload,
            )
        else:
            return Response({'error': 'Document non parsé.'}, status=400)

    outcome = _apply_to_score(request.user, result, doc_id=doc_id, file_path=file_path)
    return Response({
        'message': outcome['message'],
        'teras_signals': result.get('teras_signals', {}),
        'score': outcome.get('score'),
        'dashboard_updates': outcome.get('dashboard_updates', {}),
        'recommendations_count': outcome.get('recommendations_count', 0),
    })


# ─────────────────────────────────────────────────────────────────────────────
# HELPER — APPLICATION AU SCORE
# ─────────────────────────────────────────────────────────────────────────────

def _apply_to_score(user, parse_result: dict, *, doc_id: str, file_path: str) -> dict:
    """
    Applique les signaux du document parsé au profil TERAS de l'utilisateur.
    Met à jour les champs monthly_revenue_avg, savings_balance, etc.
    """
    signals = parse_result.get('teras_signals', {}) or {}

    try:
        doc_type = parse_result.get('doc_type', 'bank_statement')
        if doc_type == 'proof_asset':
            analysis_payload = _generate_local_user_analysis(
                user,
                parse_result,
                file_path=file_path,
                doc_id=doc_id,
            )
        else:
            analysis_payload = parse_result.get('analysis_summary') or _generate_local_user_analysis(
                user,
                parse_result,
                file_path=file_path,
                doc_id=doc_id,
            )
        if not signals and doc_type != 'proof_asset':
            return {
                'message': "Aucun signal TERAS exploitable n'a ete extrait pour ce document.",
                'dashboard_updates': analysis_payload.get('dashboard_updates', {}),
                'recommendations_count': 0,
            }
        income_signal = signals.get('income_signal', {}) or {}
        savings_signal = signals.get('savings_signal', {}) or {}
        transactions_signal = signals.get('transactions_signal', {}) or {}
        quality = parse_result.get('quality', {}) or {}
        stats = quality.get('stats', {}) or {}

        avg_income = _safe_float(income_signal.get('monthly_avg_xaf'))
        savings_avg = _safe_float(savings_signal.get('monthly_deposit_avg_xaf'))
        crm_estimated = _safe_float(signals.get('crm_estimated_xaf'))
        stability = _safe_float(income_signal.get('income_stability'))
        regularity = _safe_float(transactions_signal.get('regularity_score'))
        authenticity = _safe_float(quality.get('authenticity_score'))
        streak = int(savings_signal.get('streak_months') or 0)
        months_covered = int(stats.get('months_covered') or signals.get('months_analyzed') or 0)

        tx_prefix = f"[DOC:{doc_id}] "
        Transaction.objects.filter(user=user, description__startswith=tx_prefix).delete()
        for txn in parse_result.get('transactions', [])[:200]:
            amount = _safe_float(txn.get('amount'))
            if amount <= 0:
                continue
            Transaction.objects.create(
                user=user,
                amount=Decimal(str(round(amount, 2))),
                transaction_type='credit' if txn.get('type') == 'credit' else 'debit',
                channel=txn.get('category') or 'document',
                description=f"{tx_prefix}{txn.get('description', 'Transaction document')[:180]}",
            )

        if avg_income > 0:
            Income.objects.update_or_create(
                user=user,
                source=f"Analyse document {doc_id}",
                defaults={
                    'amount': Decimal(str(round(avg_income, 2))),
                    'is_recurring': stability >= 0.55,
                    'verified': True,
                },
            )

        asset_score_base = float(
            Asset.objects.filter(user=user, verified=True).aggregate(total=models.Sum('estimated_value'))['total']
            or Decimal('0')
        )
        if doc_type == 'proof_asset':
            synced_assets = _sync_user_assets_from_analysis(user, analysis_payload, doc_id)
            asset_score_base = float(
                Asset.objects.filter(user=user, verified=True).aggregate(total=models.Sum('estimated_value'))['total']
                or Decimal('0')
            )

        latest_score = TerasScore.objects.filter(user=user).order_by('-created_at').first()
        transactions_score = round(min(100, regularity * 70 + min(months_covered, 6) * 4 + authenticity * 18))
        savings_ratio = (savings_avg / avg_income) if avg_income > 0 else 0
        savings_score = round(min(100, min(streak, 6) * 10 + min(savings_ratio, 0.45) * 100 + authenticity * 12))
        income_score = round(min(100, min(avg_income / 750000, 1.0) * 60 + stability * 40))
        if doc_type == 'proof_asset' and not signals:
            transactions_score = latest_score.transactions_score if latest_score else 42
            savings_score = latest_score.savings_score if latest_score else 36
            income_score = latest_score.income_score if latest_score else 41
        if avg_income > 0:
            asset_cover = min(asset_score_base / max(avg_income * 12, 1), 1.4)
            assets_score = round(min(100, asset_cover / 1.4 * 100))
        else:
            assets_score = round(min(100, min(asset_score_base / 2_500_000, 1.0) * 100))
        if doc_type == 'proof_asset' and not signals:
            assets_score = max(assets_score, latest_score.assets_score if latest_score else 28)

        social_score = latest_score.social_score if latest_score else 45

        total_score = round((
            transactions_score * 0.28
            + savings_score * 0.18
            + income_score * 0.22
            + assets_score * 0.20
            + social_score * 0.12
        ) * 10)

        reason_codes = [
            f"doc:{doc_id}",
            f"auth:{round(authenticity * 100)}",
            f"months:{months_covered}",
            f"income:{round(avg_income)}",
            f"crm:{round(crm_estimated)}",
        ]
        generated_score = TerasScore.objects.create(
            user=user,
            score=max(0, min(1000, total_score)),
            transactions_score=transactions_score,
            savings_score=savings_score,
            income_score=income_score,
            assets_score=assets_score,
            social_score=social_score,
            reason_codes=reason_codes,
            model_version='document-analysis-2.0',
            source='document',
            is_simulated=False,
        )

        recommendation_specs = []
        if regularity < 0.6:
            recommendation_specs.append((
                'transactions',
                'high',
                f'Doc {doc_id}: stabiliser les entrees',
                'Maintenez des flux plus reguliers sur votre compte principal pour renforcer le pilier Transactions.',
                '+20 points',
            ))
        if stability < 0.55:
            recommendation_specs.append((
                'revenus',
                'high',
                f'Doc {doc_id}: mieux documenter les revenus',
                'Ajoutez des pieces recurrentes (bulletins, contrats, preuves de paiement) pour solidifier le pilier Revenus.',
                '+25 points',
            ))
        if savings_avg <= 0:
            recommendation_specs.append((
                'epargne',
                'medium',
                f'Doc {doc_id}: mettre en place une epargne visible',
                'Des depots d’epargne reguliers ameliorent directement la capacite de remboursement percue.',
                '+15 points',
            ))

        recommendations_count = 0
        for category, priority, title, description, impact in recommendation_specs:
            Recommendation.objects.update_or_create(
                user=user,
                title=title,
                defaults={
                    'category': category,
                    'priority': priority,
                    'description': description,
                    'impact': impact,
                    'completed': False,
                    'completed_at': None,
                },
            )
            recommendations_count += 1

        analysis_payload = {
            **analysis_payload,
            'applied_to_teras': True,
            'applied_at': timezone.now().isoformat(),
            'applied_score': {
                'id': generated_score.id,
                'score': generated_score.score,
                'level': generated_score.level_display,
                'breakdown': generated_score.breakdown,
            },
            'dashboard_updates': {
                **analysis_payload.get('dashboard_updates', {}),
                'monthly_income_xaf': round(avg_income),
                'monthly_savings_xaf': round(savings_avg),
                'crm_estimated_xaf': round(crm_estimated),
                'transactions_synced': len(parse_result.get('transactions', [])[:200]),
                'estimated_asset_value_xaf': round(asset_score_base),
                'asset_items_count': len(analysis_payload.get('detected_assets') or []),
                'assets_synced_count': synced_assets.get('assets_count', 0) if doc_type == 'proof_asset' else 0,
            },
        }
        parse_result['analysis_summary'] = analysis_payload
        parse_result['analysis_text'] = analysis_payload.get('analysis_text', '')
        parse_result['generated_score'] = analysis_payload['applied_score']
        parse_result['processed_at'] = timezone.now().isoformat()
        _save_result(file_path, parse_result)
        _sync_user_document_record(
            user,
            file_path,
            doc_id,
            doc_type=parse_result.get('doc_type', 'bank_statement'),
            parse_result=parse_result,
            analysis_payload=analysis_payload,
            generated_score=generated_score,
        )

        return {
            'message': (
                f"Analyse appliquee au moteur TERAS. Nouveau score : {generated_score.score}/1000 "
                f"({generated_score.level_display}). Dashboard mis a jour avec revenu estime a {round(avg_income):,} FCFA."
            ),
            'score': {
                'id': generated_score.id,
                'value': generated_score.score,
                'level': generated_score.level_display,
                'breakdown': generated_score.breakdown,
            },
            'dashboard_updates': analysis_payload['dashboard_updates'],
            'recommendations_count': recommendations_count,
        }

    except Exception as e:
        logger.error(f"Erreur application score : {e}")
        return {
            'message': f"Erreur application : {e}",
            'dashboard_updates': {},
            'recommendations_count': 0,
        }
